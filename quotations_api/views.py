from datetime import datetime
import json
from django.db.models import Q
from django.shortcuts import get_object_or_404
from rest_framework import status, viewsets, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from .models import Quotation, Payment, Delivery, Other, QuotationItem, LastQuotedPrice
from admin_api.models import Customer, CustomerContact, Inventory
from .serializers import (
    QuotationSerializer, QuotationCreateUpdateSerializer, CustomerListSerializer,
    PaymentSerializer, DeliverySerializer, OtherSerializer, CustomerContactSerializer,
    QuotationStatusUpdateSerializer, LastQuotedPriceSerializer
)
from django.http import HttpResponse, FileResponse
from .pdf_template import generate_quotation_pdf
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from rest_framework.decorators import action

class QuotationView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        # If pk is provided, return a single quotation with all related data
        if pk:
            quotation = get_object_or_404(Quotation, pk=pk)
            serializer = QuotationSerializer(quotation)
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get search parameters for specific fields
        quote_number_search = request.query_params.get('quote_number', '')
        status = request.query_params.get('status', '')
        customer = request.query_params.get('customer', '')
        date_from = request.query_params.get('date_from', '')
        date_to = request.query_params.get('date_to', '')
        
        # Get general search parameter
        general_search = request.query_params.get('search', '')
        
        # Get sorting parameters
        sort_by = request.query_params.get('sort_by', '-date')
        sort_direction = request.query_params.get('sort_direction', 'asc')
        
        # Query quotations
        quotations = Quotation.objects.all()

        # Apply field-specific search filters
        if quote_number_search:
            quotations = quotations.filter(quote_number__icontains=quote_number_search)
        
        if status and status in dict(Quotation.STATUS_CHOICES):
            quotations = quotations.filter(status=status)
            
        if customer:
            quotations = quotations.filter(customer__name__icontains=customer)
            
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                quotations = quotations.filter(date__gte=date_from_obj)
            except ValueError:
                pass
                
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                quotations = quotations.filter(date__lte=date_to_obj)
            except ValueError:
                pass

        # Apply general search filter if no specific filters are provided
        if general_search and not any([quote_number_search, status, customer, date_from, date_to]):
            quotations = quotations.filter(
                Q(quote_number__icontains=general_search) |
                Q(customer__name__icontains=general_search) |
                Q(sales_agents__agent_name__icontains=general_search)
            ).distinct()

        # Apply sorting
        sort_prefix = '-' if sort_direction == 'desc' else ''
        sort_field = sort_by.lstrip('-')
        sort_order = f"{sort_prefix}{sort_field}"
        quotations = quotations.order_by(sort_order)
        
        # Pagination
        page = self.paginate_queryset(quotations, request)
        if page is not None:
            serializer = QuotationSerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    },
                    'currency_options': ['USD', 'EURO', 'RMB', 'PHP'],
                    'status_options': ['draft', 'for_approval', 'approved', 'expired'],
                }
            })

        # Fallback if pagination fails
        serializer = QuotationSerializer(quotations, many=True)
        return Response({
            'success': True,
            'data': serializer.data,
            'meta': {
                'currency_options': ['USD', 'EURO', 'RMB', 'PHP'],
                'status_options': ['draft', 'for_approval', 'approved', 'expired'],
            }
        })

    def post(self, request):
        try:
            # Extract the JSON data from the 'data' field
            if 'data' in request.data:
                json_data = json.loads(request.data['data'])
                
                # Process attachments if any
                if 'attachments' in json_data and json_data['attachments']:
                    attachments_data = []
                    for i, attachment in enumerate(json_data['attachments']):
                        file_key = f'attachments[{i}][file]'
                        if file_key in request.data:
                            attachment_data = {
                                'file': request.data[file_key],
                                'filename': attachment.get('filename', '')
                            }
                            attachments_data.append(attachment_data)
                    
                    # Replace the attachments in json_data with the processed ones
                    json_data['attachments'] = attachments_data
                
                serializer = QuotationCreateUpdateSerializer(data=json_data, context={'request': request})
                
                if serializer.is_valid():
                    serializer.save()
                    return Response({
                        'success': True,
                        'data': serializer.data
                    }, status=status.HTTP_201_CREATED)
                else:
                    # Format validation errors
                    error_messages = {}
                    for field, errors in serializer.errors.items():
                        error_messages[field] = errors[0] if isinstance(errors, list) else errors
                    return Response({
                        'success': False,
                        'errors': error_messages
                    }, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({
                    'success': False,
                    'errors': {'detail': 'No data provided'}
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk):
        """Update a quotation"""
        try:
            quotation = get_object_or_404(Quotation, pk=pk)
            
            # Parse the JSON data
            data = {}
            if 'data' in request.data:
                try:
                    data = json.loads(request.data['data'])
                except json.JSONDecodeError as e:
                    return Response({
                        'success': False,
                        'errors': {'detail': f'Invalid JSON data: {str(e)}'}
                    }, status=status.HTTP_400_BAD_REQUEST)
            else:
                # If no 'data' field, use the request.data directly
                data = request.data
            
            # Handle file uploads
            files = request.FILES.getlist('files') if 'files' in request.FILES else []
            
            # Create serializer with the data
            serializer = QuotationCreateUpdateSerializer(
                quotation, 
                data=data,
                partial=True,  # Allow partial updates
                context={'request': request, 'files': files}
            )
            
            if serializer.is_valid():
                updated_quotation = serializer.save()
                
                # Return the updated quotation
                return Response({
                    'success': True,
                    'data': QuotationSerializer(updated_quotation).data
                })
            else:
                return Response({
                    'success': False,
                    'errors': serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self, request, pk):
        quotation = get_object_or_404(Quotation, pk=pk)
        quotation.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class CustomerListView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Get only active customers
        customers = Customer.objects.filter(status='active')
        serializer = CustomerListSerializer(customers, many=True)
        
        return Response({
            'success': True,
            'data': serializer.data
        })

class PaymentView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, pk=None):
        if pk:
            payment = get_object_or_404(Payment, pk=pk)
            serializer = PaymentSerializer(payment)
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get search parameter
        search = request.query_params.get('search', '')
        
        # Query payments
        payments = Payment.objects.all()
        
        # Apply search filter
        if search:
            payments = payments.filter(text__icontains=search)
        
        # Order by most recent
        payments = payments.order_by('-created_on')
        
        # Pagination
        page = self.paginate_queryset(payments, request)
        if page is not None:
            serializer = PaymentSerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })
        
        # Fallback if pagination fails
        serializer = PaymentSerializer(payments, many=True)
        return Response({
            'success': True,
            'data': serializer.data
        })
    
    def post(self, request):
        serializer = PaymentSerializer(data=request.data)
        if serializer.is_valid():
            # Set the created_by field
            payment = serializer.save(created_by=request.user)
            return Response({
                'success': True,
                'data': PaymentSerializer(payment).data
            }, status=status.HTTP_201_CREATED)
        else:
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        payment = get_object_or_404(Payment, pk=pk)
        payment.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class DeliveryView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, pk=None):
        if pk:
            delivery = get_object_or_404(Delivery, pk=pk)
            serializer = DeliverySerializer(delivery)
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get search parameter
        search = request.query_params.get('search', '')
        
        # Query deliveries
        deliveries = Delivery.objects.all()
        
        # Apply search filter
        if search:
            deliveries = deliveries.filter(text__icontains=search)
        
        # Order by most recent
        deliveries = deliveries.order_by('-created_on')
        
        # Pagination
        page = self.paginate_queryset(deliveries, request)
        if page is not None:
            serializer = DeliverySerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })
        
        # Fallback if pagination fails
        serializer = DeliverySerializer(deliveries, many=True)
        return Response({
            'success': True,
            'data': serializer.data
        })
    
    def post(self, request):
        serializer = DeliverySerializer(data=request.data)
        if serializer.is_valid():
            # Set the created_by field
            delivery = serializer.save(created_by=request.user)
            return Response({
                'success': True,
                'data': DeliverySerializer(delivery).data
            }, status=status.HTTP_201_CREATED)
        else:
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        delivery = get_object_or_404(Delivery, pk=pk)
        delivery.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class OtherView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, pk=None):
        if pk:
            other = get_object_or_404(Other, pk=pk)
            serializer = OtherSerializer(other)
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get search parameter
        search = request.query_params.get('search', '')
        
        # Query others
        others = Other.objects.all()
        
        # Apply search filter
        if search:
            others = others.filter(text__icontains=search)
        
        # Order by most recent
        others = others.order_by('-created_on')
        
        # Pagination
        page = self.paginate_queryset(others, request)
        if page is not None:
            serializer = OtherSerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })
        
        # Fallback if pagination fails
        serializer = OtherSerializer(others, many=True)
        return Response({
            'success': True,
            'data': serializer.data
        })
    
    def post(self, request):
        serializer = OtherSerializer(data=request.data)
        if serializer.is_valid():
            # Set the created_by field
            other = serializer.save(created_by=request.user)
            return Response({
                'success': True,
                'data': OtherSerializer(other).data
            }, status=status.HTTP_201_CREATED)
        else:
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        other = get_object_or_404(Other, pk=pk)
        other.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class CustomerContactListView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        customer_id = request.query_params.get('customer_id')
        
        if not customer_id:
            return Response({
                'success': False,
                'errors': {'detail': 'Customer ID is required'}
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get contacts for the specified customer
        contacts = CustomerContact.objects.filter(customer_id=customer_id)
        
        # Apply search if provided
        search = request.query_params.get('search', '')
        if search:
            contacts = contacts.filter(contact_person__icontains=search)
        
        # Paginate results
        page = self.paginate_queryset(contacts, request)
        if page is not None:
            serializer = CustomerContactSerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })
        
        # Fallback if pagination fails
        serializer = CustomerContactSerializer(contacts, many=True)
        return Response({
            'success': True,
            'data': serializer.data
        })
    
    def post(self, request):
        """Add a new contact to the customer's contacts"""
        serializer = CustomerContactSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                'success': True,
                'data': serializer.data
            }, status=status.HTTP_201_CREATED)
        else:
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

class QuotationPDFView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, pk):
        """
        Generate and download a PDF for the specified quotation
        """
        try:
            # Get the quotation
            quotation = get_object_or_404(Quotation, pk=pk)
            
            # Generate the PDF
            pdf_buffer = generate_quotation_pdf(quotation)
            
            # Create the HTTP response with PDF content
            response = HttpResponse(pdf_buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{quotation.quote_number}.pdf"'
            
            return response
        except Exception as e:
            import traceback
            print(f"PDF generation error: {str(e)}")
            print(traceback.format_exc())
            return Response(
                {'success': False, 'errors': {'detail': str(e)}},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class QuotationItemsTemplateView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, pk=None):
        """Download an Excel template for bulk uploading quotation items"""
        # Verify the quotation exists
        quotation = get_object_or_404(Quotation, pk=pk)
        
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Items Template"
        
        # Add headers
        headers = ['item_code', 'quantity']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
        
        # Add example row
        ws['A2'] = 'ABC123'
        ws['B2'] = 1
        
        # Save to a BytesIO object
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="quotation_{quotation.quote_number}_items_template.xlsx"'
        
        return response

class QuotationItemsUploadView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request, pk=None):
        """Upload an Excel file with item codes and quantities to add to the quotation"""
        try:
            quotation = get_object_or_404(Quotation, pk=pk)
            
            if 'file' not in request.FILES:
                return Response(
                    {'success': False, 'errors': 'No file uploaded'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            file = request.FILES['file']
            
            # Check file extension
            if not (file.name.endswith('.xlsx') or file.name.endswith('.xls')):
                return Response(
                    {'success': False, 'errors': 'File must be an Excel file (.xlsx or .xls)'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Process the Excel file
            try:
                # Load the workbook
                wb = load_workbook(filename=io.BytesIO(file.read()))
                ws = wb.active
                
                # Get headers from the first row
                headers = [cell.value for cell in ws[1]]
                
                # Validate required columns
                required_columns = ['item_code', 'quantity']
                for col in required_columns:
                    if col not in headers:
                        return Response(
                            {'success': False, 'errors': f'Missing required column: {col}'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                
                # Get column indices
                item_code_idx = headers.index('item_code')
                quantity_idx = headers.index('quantity')
                
                # Process each row
                results = {
                    'success': True,
                    'added': 0,
                    'errors': [],
                    'total_rows': ws.max_row - 1  # Subtract header row
                }
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    # Validate item_code
                    item_code = str(row[item_code_idx] or '').strip()
                    if not item_code:
                        results['errors'].append(f'Line {row_idx}: Item code is empty')
                        continue
                    
                    # Validate quantity
                    try:
                        quantity_value = row[quantity_idx]
                        if quantity_value is None:
                            results['errors'].append(f'Line {row_idx}: Quantity is empty')
                            continue
                            
                        quantity = int(float(quantity_value))
                        if quantity <= 0:
                            results['errors'].append(f'Line {row_idx}: Quantity must be a positive number')
                            continue
                    except (ValueError, TypeError):
                        results['errors'].append(f'Line {row_idx}: Invalid quantity format')
                        continue
                    
                    # Find inventory item
                    try:
                        inventory = Inventory.objects.get(item_code=item_code)
                    except Inventory.DoesNotExist:
                        results['errors'].append(f'Line {row_idx}: Item code "{item_code}" not found')
                        continue
                    
                    # Create quotation item
                    try:
                        # Check if item already exists in this quotation
                        existing_item = QuotationItem.objects.filter(quotation=quotation, inventory=inventory).first()
                        
                        if existing_item:
                            # Update quantity if item already exists
                            existing_item.quantity = quantity
                            existing_item.save()
                        else:
                            # Create new item
                            QuotationItem.objects.create(
                                quotation=quotation,
                                inventory=inventory,
                                quantity=quantity,
                                wholesale_price=inventory.wholesale_price,
                                unit=inventory.unit,
                                external_description=inventory.external_description
                            )
                        
                        results['added'] += 1
                    except Exception as e:
                        results['errors'].append(f'Line {row_idx}: Failed to add item - {str(e)}')
                
                # Update quotation total amount
                quotation_items = quotation.items.all()
                total_amount = 0
                for item in quotation_items:
                    if item.total_selling is not None:
                        total_amount += item.total_selling
                
                quotation.total_amount = total_amount
                quotation.save()
                
                return Response(results)
                
            except Exception as e:
                import traceback
                print(traceback.format_exc())
                return Response(
                    {'success': False, 'errors': f'Error processing Excel file: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return Response(
                {'success': False, 'errors': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class QuotationViewSet(viewsets.ModelViewSet):
    queryset = Quotation.objects.all()
    serializer_class = QuotationSerializer
    permission_classes = [IsAuthenticated]
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, last_modified_by=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(last_modified_by=self.request.user)
    
    @action(detail=True, methods=['patch'], permission_classes=[IsAuthenticated])
    def update_status(self, request, pk=None):
        quotation = self.get_object()
        serializer = QuotationStatusUpdateSerializer(
            quotation, 
            data=request.data,
            context={'request': request}
        )
        
        if serializer.is_valid():
            serializer.save()
            # Return the full quotation data with updated status
            return Response(QuotationSerializer(quotation).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LastQuotedPriceViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = LastQuotedPrice.objects.all()
    serializer_class = LastQuotedPriceSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = LastQuotedPrice.objects.all()
        
        # Filter by customer if provided
        customer_id = self.request.query_params.get('customer')
        if customer_id:
            queryset = queryset.filter(customer_id=customer_id)
            
        # Filter by inventory if provided
        inventory_id = self.request.query_params.get('inventory')
        if inventory_id:
            queryset = queryset.filter(inventory_id=inventory_id)
            
        return queryset

class QuotationStatusView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request, pk):
        """Update the status of a quotation"""
        quotation = get_object_or_404(Quotation, pk=pk)
        new_status = request.data.get('status')
        
        # Validate the requested status transition
        if not new_status or new_status not in dict(Quotation.STATUS_CHOICES):
            return Response({
                'success': False,
                'errors': {'status': 'Invalid status value'}
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if user has permission for this status change
        current_status = quotation.status
        
        # Any user can send a draft quotation for approval
        if current_status == 'draft' and new_status == 'for_approval':
            pass  # Allow this transition
        # Only admin/supervisor can approve or reject
        elif current_status == 'for_approval' and new_status in ['approved', 'rejected']:
            # Check if user is admin or supervisor
            if not (request.user.is_staff or request.user.groups.filter(name='Supervisor').exists()):
                return Response({
                    'success': False,
                    'errors': {'detail': 'You do not have permission to approve or reject quotations'}
                }, status=status.HTTP_403_FORBIDDEN)
        else:
            return Response({
                'success': False,
                'errors': {'status': f'Cannot change status from {current_status} to {new_status}'}
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Update the quotation status
        quotation.status = new_status
        quotation.last_modified_by = request.user
        quotation.save()
        
        # If approved, save the last quoted prices
        if new_status == 'approved':
            self._save_last_quoted_prices(quotation)
        
        # Return the updated quotation
        return Response({
            'success': True,
            'data': QuotationSerializer(quotation).data
        })
    
    def _save_last_quoted_prices(self, quotation):
        """Save the last quoted prices for all items in the quotation"""
        for item in quotation.items.all():
            # Update or create LastQuotedPrice entry
            LastQuotedPrice.objects.update_or_create(
                inventory=item.inventory,
                customer=quotation.customer,
                defaults={
                    'price': item.wholesale_price,
                    'quotation': quotation,
                }
            )

class LastQuotedPriceView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Get last quoted prices with optional filtering"""
        # Get filter parameters
        customer_id = request.query_params.get('customer_id')
        inventory_id = request.query_params.get('inventory_id')
        
        # Start with all records
        queryset = LastQuotedPrice.objects.all()
        
        # Apply filters if provided
        if customer_id:
            queryset = queryset.filter(customer_id=customer_id)
        
        if inventory_id:
            queryset = queryset.filter(inventory_id=inventory_id)
        
        # Order by most recent
        queryset = queryset.order_by('-quoted_at')
        
        # Paginate results
        page = self.paginate_queryset(queryset, request)
        if page is not None:
            serializer = LastQuotedPriceSerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })
        
        # Fallback if pagination fails
        serializer = LastQuotedPriceSerializer(queryset, many=True)
        return Response({
            'success': True,
            'data': serializer.data
        })