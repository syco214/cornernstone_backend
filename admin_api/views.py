from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework_simplejwt.tokens import RefreshToken
from .serializers import LoginSerializer
from django.shortcuts import get_object_or_404
from django.db.models import Q
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from rest_framework.parsers import MultiPartParser, FormParser

from .models import USER_ACCESS_OPTIONS, USER_ROLE_OPTIONS, CustomUser, Brand, Category, Warehouse, Supplier, ParentCompany, Customer, Broker, Forwarder, Inventory, ADMIN_ACCESS_OPTIONS
from .serializers import UserSerializer, SidebarUserSerializer, BrandSerializer, CategorySerializer, CategoryTreeSerializer, WarehouseSerializer, WarehouseCreateUpdateSerializer, SupplierSerializer, SupplierCreateUpdateSerializer, ParentCompanySerializer, ParentCompanyPaymentTermSerializer, ParentCompanyCreateUpdateSerializer, CustomerSerializer, CustomerCreateUpdateSerializer, BrokerSerializer, BrokerCreateUpdateSerializer, ForwarderSerializer, ForwarderCreateUpdateSerializer, InventorySerializer

# Create your views here.

class LoginView(APIView):
    permission_classes = [AllowAny]
    serializer_class = LoginSerializer

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        try:
            serializer.is_valid(raise_exception=True)
            user = serializer.validated_data
            refresh = RefreshToken.for_user(user)

            return Response({
                'success': True,
                'data': {
                    'token': str(refresh.access_token),
                    'refresh': str(refresh),
                    'user': {
                        'username': user.username,
                        'id': user.id,
                    }
                }
            })
        except Exception as e:
            return Response({
                'success': False,
                'error': 'Invalid credentials'
            }, status=status.HTTP_401_UNAUTHORIZED)

class SidebarView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        serializer = SidebarUserSerializer(user)
        
        return Response({
            'success': True,
            'data': serializer.data
        }, status=status.HTTP_200_OK)
    
class UserView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        # If pk is provided, return a single user
        if pk:
            user = get_object_or_404(CustomUser, pk=pk)
            serializer = UserSerializer(user)
            return Response({
                'success': True,
                'data': serializer.data,
                'meta': {
                    'user_access_options': USER_ACCESS_OPTIONS,
                    'user_role_options': USER_ROLE_OPTIONS,
                    'admin_access_options': ADMIN_ACCESS_OPTIONS
                }
            })
        
        # Otherwise, return a list of users
        # Get search parameters for specific fields
        username_search = request.query_params.get('username', '')
        first_name_search = request.query_params.get('first_name', '')
        last_name_search = request.query_params.get('last_name', '')
        role_search = request.query_params.get('role', '')
        
        # Get general search parameter
        general_search = request.query_params.get('search', '')
        
        # Get sorting parameters
        sort_by = request.query_params.get('sort_by', 'id')
        sort_direction = request.query_params.get('sort_direction', 'asc')

        # Query users
        users = CustomUser.objects.all()

        # Apply field-specific search filters
        if username_search:
            users = users.filter(username__icontains=username_search)
        
        if first_name_search:
            users = users.filter(first_name__icontains=first_name_search)
        
        if last_name_search:
            users = users.filter(last_name__icontains=last_name_search)
            
        if role_search:
            users = users.filter(role__iexact=role_search)

        # Apply general search filter if no specific filters are provided
        if general_search and not any([username_search, first_name_search, last_name_search, role_search]):
            users = users.filter(
                Q(first_name__icontains=general_search) |
                Q(last_name__icontains=general_search) |
                Q(username__icontains=general_search) |
                Q(role__icontains=general_search)
            )

        # Apply sorting
        sort_prefix = '-' if sort_direction == 'desc' else ''
        users = users.order_by(f'{sort_prefix}{sort_by}')

        # Pagination is handled by the mixin
        page = self.paginate_queryset(users, request)
        if page is not None:
            serializer = UserSerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            # Restructure to match our API format
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    },
                    'user_access_options': USER_ACCESS_OPTIONS,
                    'user_role_options': USER_ROLE_OPTIONS,
                    'admin_access_options': ADMIN_ACCESS_OPTIONS
                }
            })

        # Fallback if pagination fails
        serializer = UserSerializer(users, many=True)
        return Response({
            'success': True,
            'data': serializer.data,
            'meta': {
                'user_access_options': USER_ACCESS_OPTIONS,
                'user_role_options': USER_ROLE_OPTIONS,
                'admin_access_options': ADMIN_ACCESS_OPTIONS
            }
        })

    def post(self, request):
        serializer = UserSerializer(data=request.data)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data,
                    'meta': {
                        'user_access_options': USER_ACCESS_OPTIONS,
                        'user_role_options': USER_ROLE_OPTIONS,
                        'admin_access_options': ADMIN_ACCESS_OPTIONS
                    }
                }, status=status.HTTP_201_CREATED)
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    # Convert list of errors to single string
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk):
        user = get_object_or_404(CustomUser, pk=pk)
        serializer = UserSerializer(user, data=request.data, partial=True)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data,
                    'meta': {
                        'user_access_options': USER_ACCESS_OPTIONS,
                        'user_role_options': USER_ROLE_OPTIONS,
                        'admin_access_options': ADMIN_ACCESS_OPTIONS
                    }
                })
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        user = get_object_or_404(CustomUser, pk=pk)
        user.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class BrandView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        # If pk is provided, return a single brand
        if pk:
            brand = get_object_or_404(Brand, pk=pk)
            serializer = BrandSerializer(brand)
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Otherwise, return a list of brands
        # Get search parameters for specific fields
        id_search = request.query_params.get('id', '')
        name_search = request.query_params.get('name', '')
        made_in_search = request.query_params.get('made_in', '')
        show_made_in = request.query_params.get('show_made_in', '')
        
        # Get general search parameter
        general_search = request.query_params.get('search', '')
        
        # Get sorting parameters
        sort_by = request.query_params.get('sort_by', 'name')
        sort_direction = request.query_params.get('sort_direction', 'asc')

        # Query brands
        brands = Brand.objects.all()

        # Apply field-specific search filters
        if id_search:
            try:
                brands = brands.filter(id=int(id_search))
            except ValueError:
                # If id_search is not a valid integer, return empty queryset
                brands = Brand.objects.none()
        
        if name_search:
            brands = brands.filter(name__icontains=name_search)
        
        if made_in_search:
            brands = brands.filter(made_in__icontains=made_in_search)
        
        if show_made_in:
            show_made_in_bool = show_made_in.lower() in ['true', '1', 'yes']
            brands = brands.filter(show_made_in=show_made_in_bool)

        # Apply general search filter if no specific filters are provided
        if general_search and not any([id_search, name_search, made_in_search, show_made_in]):
            brands = brands.filter(
                Q(name__icontains=general_search) |
                Q(made_in__icontains=general_search) |
                Q(remarks__icontains=general_search)
            )

        # Apply sorting
        sort_prefix = '-' if sort_direction == 'desc' else ''
        brands = brands.order_by(f'{sort_prefix}{sort_by}')

        # Pagination is handled by the mixin
        page = self.paginate_queryset(brands, request)
        if page is not None:
            serializer = BrandSerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            # Restructure to match our API format
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })

        # Fallback if pagination fails
        serializer = BrandSerializer(brands, many=True)
        return Response({
            'success': True,
            'data': serializer.data
        })

    def post(self, request):
        serializer = BrandSerializer(data=request.data)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    # Convert list of errors to single string
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk):
        brand = get_object_or_404(Brand, pk=pk)
        serializer = BrandSerializer(brand, data=request.data, partial=True)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                })
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        brand = get_object_or_404(Brand, pk=pk)
        brand.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class CategoryView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        # If pk is provided, return a single category
        if pk:
            category = get_object_or_404(Category, pk=pk)
            serializer = CategorySerializer(category)
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get query parameters
        search = request.query_params.get('search', '')
        sort_by = request.query_params.get('sort_by', 'name')
        sort_direction = request.query_params.get('sort_direction', 'asc')
        tree_view = request.query_params.get('tree', 'false').lower() == 'true'
        parent_id = request.query_params.get('parent', None)
        
        # Query categories
        if parent_id:
            if parent_id == 'root':
                # Get only root categories (no parent)
                categories = Category.objects.filter(parent=None)
            else:
                # Get categories under specific parent
                categories = Category.objects.filter(parent_id=parent_id)
        else:
            # Get all categories
            categories = Category.objects.all()

        # Apply search filter
        if search:
            categories = categories.filter(
                Q(name__icontains=search)
            )

        # Apply sorting
        sort_prefix = '-' if sort_direction == 'desc' else ''
        categories = categories.order_by(f'{sort_prefix}{sort_by}')

        # Handle tree view (only for root level when requested)
        if tree_view and not parent_id and not search:
            root_categories = Category.objects.filter(parent=None)
            serializer = CategoryTreeSerializer(root_categories, many=True)
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Pagination for flat view
        page = self.paginate_queryset(categories, request)
        if page is not None:
            serializer = CategorySerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })

        # Fallback if pagination fails
        serializer = CategorySerializer(categories, many=True)
        return Response({
            'success': True,
            'data': serializer.data
        })

    def post(self, request):
        serializer = CategorySerializer(data=request.data)
        try:
            if serializer.is_valid():
                # Check for circular references
                parent_id = request.data.get('parent')
                if parent_id:
                    parent = get_object_or_404(Category, pk=parent_id)
                    # Prevent a category from being its own ancestor
                    current_parent = parent
                    while current_parent:
                        if str(current_parent.id) == request.data.get('id', '-1'):
                            return Response({
                                'success': False,
                                'errors': {'parent': 'Circular reference detected. A category cannot be its own ancestor.'}
                            }, status=status.HTTP_400_BAD_REQUEST)
                        current_parent = current_parent.parent
                
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk):
        category = get_object_or_404(Category, pk=pk)
        serializer = CategorySerializer(category, data=request.data, partial=True)
        try:
            if serializer.is_valid():
                # Check for circular references
                parent_id = request.data.get('parent')
                if parent_id:
                    # Prevent setting itself as parent
                    if str(parent_id) == str(pk):
                        return Response({
                            'success': False,
                            'errors': {'parent': 'A category cannot be its own parent.'}
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    # Prevent setting one of its descendants as parent
                    descendants = self._get_all_descendants(category)
                    if int(parent_id) in [desc.id for desc in descendants]:
                        return Response({
                            'success': False,
                            'errors': {'parent': 'Cannot set a descendant as parent (circular reference).'}
                        }, status=status.HTTP_400_BAD_REQUEST)
                
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                })
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        category = get_object_or_404(Category, pk=pk)
        category.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)
    
    def _get_all_descendants(self, category):
        """Helper method to get all descendants of a category"""
        descendants = []
        children = Category.objects.filter(parent=category)
        
        for child in children:
            descendants.append(child)
            descendants.extend(self._get_all_descendants(child))
        
        return descendants

class CategoryChildrenView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, parent_id):
        # Get subcategories for a specific parent
        categories = Category.objects.filter(parent_id=parent_id)
        serializer = CategorySerializer(categories, many=True)
        return Response({
            'success': True,
            'data': serializer.data
        })
    
class WarehouseView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        # If pk is provided, return a single warehouse with its shelves
        if pk:
            warehouse = get_object_or_404(Warehouse, pk=pk)
            serializer = WarehouseSerializer(warehouse)
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get search parameters for specific fields
        id_search = request.query_params.get('id', '')
        name_search = request.query_params.get('name', '')
        city_search = request.query_params.get('city', '')
        
        # Get general search parameter
        general_search = request.query_params.get('search', '')
        
        # Get sorting parameters
        sort_by = request.query_params.get('sort_by', 'name')
        sort_direction = request.query_params.get('sort_direction', 'asc')
        
        # Query warehouses
        warehouses = Warehouse.objects.all()

        # Apply field-specific search filters
        if id_search:
            try:
                warehouses = warehouses.filter(id=int(id_search))
            except ValueError:
                # If id_search is not a valid integer, return empty queryset
                warehouses = Warehouse.objects.none()
        
        if name_search:
            warehouses = warehouses.filter(name__icontains=name_search)
        
        if city_search:
            warehouses = warehouses.filter(city__icontains=city_search)

        # Apply general search filter if no specific filters are provided
        if general_search and not any([id_search, name_search, city_search]):
            warehouses = warehouses.filter(
                Q(name__icontains=general_search) |
                Q(city__icontains=general_search) |
                Q(address__icontains=general_search)
            )

        # Apply sorting
        sort_prefix = '-' if sort_direction == 'desc' else ''
        warehouses = warehouses.order_by(f'{sort_prefix}{sort_by}')
        
        # Pagination
        page = self.paginate_queryset(warehouses, request)
        if page is not None:
            serializer = WarehouseSerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })

        # Fallback if pagination fails
        serializer = WarehouseSerializer(warehouses, many=True)
        return Response({
            'success': True,
            'data': serializer.data
        })

    def post(self, request):
        serializer = WarehouseCreateUpdateSerializer(data=request.data)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk):
        warehouse = get_object_or_404(Warehouse, pk=pk)
        serializer = WarehouseCreateUpdateSerializer(warehouse, data=request.data, partial=True)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                })
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        warehouse = get_object_or_404(Warehouse, pk=pk)
        warehouse.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class SupplierView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        # If pk is provided, return a single supplier with all related data
        if pk:
            supplier = get_object_or_404(Supplier, pk=pk)
            serializer = SupplierSerializer(supplier)
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get search parameters for specific fields
        name_search = request.query_params.get('name', '')
        supplier_type = request.query_params.get('supplier_type', '')
        
        # Get general search parameter
        general_search = request.query_params.get('search', '')
        
        # Get sorting parameters
        sort_by = request.query_params.get('sort_by', 'name')
        sort_direction = request.query_params.get('sort_direction', 'asc')
        
        # Query suppliers
        suppliers = Supplier.objects.all()

        # Apply field-specific search filters
        if name_search:
            suppliers = suppliers.filter(name__icontains=name_search)
        
        # Filter by supplier type if provided
        if supplier_type and supplier_type in dict(Supplier.SUPPLIER_TYPES):
            suppliers = suppliers.filter(supplier_type=supplier_type)

        # Apply general search filter if no specific filters are provided
        if general_search and not any([name_search, supplier_type]):
            suppliers = suppliers.filter(
                Q(name__icontains=general_search) |
                Q(email__icontains=general_search)
            )

        # Apply sorting
        sort_prefix = '-' if sort_direction == 'desc' else ''
        suppliers = suppliers.order_by(f'{sort_prefix}{sort_by}')
        
        # Pagination
        page = self.paginate_queryset(suppliers, request)
        if page is not None:
            serializer = SupplierSerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    },
                    'currency_options': ['USD', 'EURO', 'RMB', 'PHP'],
                    'supplier_type_options': ['local', 'foreign'],
                }
            })

        # Fallback if pagination fails
        serializer = SupplierSerializer(suppliers, many=True)
        return Response({
            'success': True,
            'data': serializer.data,
            'meta': {
                'currency_options': ['USD', 'EURO', 'RMB', 'PHP'],
                'supplier_type_options': ['local', 'foreign'],
            }
        })

    def post(self, request):
        serializer = SupplierCreateUpdateSerializer(data=request.data)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk):
        supplier = get_object_or_404(Supplier, pk=pk)
        serializer = SupplierCreateUpdateSerializer(supplier, data=request.data, partial=True)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                })
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        supplier = get_object_or_404(Supplier, pk=pk)
        supplier.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class ParentCompanyView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        # If pk is provided, return a single parent company with its payment terms and customers
        if pk:
            parent_company = get_object_or_404(ParentCompany, pk=pk)
            serializer = ParentCompanySerializer(parent_company)
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get search parameters for specific fields
        id_search = request.query_params.get('id', '')
        name_search = request.query_params.get('name', '')
        
        # Get general search parameter
        general_search = request.query_params.get('search', '')
        
        # Get sorting parameters
        sort_by = request.query_params.get('sort_by', 'name')
        sort_direction = request.query_params.get('sort_direction', 'asc')
        
        # Query parent companies
        parent_companies = ParentCompany.objects.all()

        # Apply field-specific search filters
        if id_search:
            try:
                parent_companies = parent_companies.filter(id=int(id_search))
            except ValueError:
                # If id_search is not a valid integer, return empty queryset
                parent_companies = ParentCompany.objects.none()
        
        if name_search:
            parent_companies = parent_companies.filter(name__icontains=name_search)

        # Apply general search filter if no specific filters are provided
        if general_search and not any([id_search, name_search]):
            parent_companies = parent_companies.filter(
                Q(name__icontains=general_search)
            )

        # Apply sorting
        sort_prefix = '-' if sort_direction == 'desc' else ''
        parent_companies = parent_companies.order_by(f'{sort_prefix}{sort_by}')
        
        # Pagination
        page = self.paginate_queryset(parent_companies, request)
        if page is not None:
            serializer = ParentCompanySerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })

        # Fallback if pagination fails
        serializer = ParentCompanySerializer(parent_companies, many=True)
        return Response({
            'success': True,
            'data': serializer.data
        })

    def post(self, request):
        serializer = ParentCompanyCreateUpdateSerializer(data=request.data)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk):
        parent_company = get_object_or_404(ParentCompany, pk=pk)
        serializer = ParentCompanyCreateUpdateSerializer(parent_company, data=request.data, partial=True)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                })
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        parent_company = get_object_or_404(ParentCompany, pk=pk)
        parent_company.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class CustomerView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        # If pk is provided, return a single customer with all related data
        if pk:
            customer = get_object_or_404(Customer, pk=pk)
            serializer = CustomerSerializer(customer)
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get search parameters for specific fields
        name_search = request.query_params.get('name', '')
        registered_name_search = request.query_params.get('registered_name', '')
        parent_company_id = request.query_params.get('parent_company_id', '')
        parent_company_name = request.query_params.get('parent_company_name', '')
        
        # Get other filter parameters
        status = request.query_params.get('status', '')
        
        # Get general search parameter
        general_search = request.query_params.get('search', '')
        
        # Get sorting parameters
        sort_by = request.query_params.get('sort_by', 'name')
        sort_direction = request.query_params.get('sort_direction', 'asc')
        
        # Query customers
        customers = Customer.objects.all()

        # Apply field-specific search filters
        if name_search:
            customers = customers.filter(name__icontains=name_search)
        
        if registered_name_search:
            customers = customers.filter(registered_name__icontains=registered_name_search)
        
        # Filter by parent company ID if provided
        if parent_company_id:
            try:
                parent_company_id = int(parent_company_id)
                customers = customers.filter(parent_company_id=parent_company_id)
            except ValueError:
                pass
        
        # Filter by parent company name if provided
        if parent_company_name:
            customers = customers.filter(parent_company__name__icontains=parent_company_name)
        
        # Filter by status if provided
        if status and status in dict(Customer.STATUS_CHOICES):
            customers = customers.filter(status=status)

        # Apply general search filter if no specific filters are provided
        if general_search and not any([name_search, registered_name_search, parent_company_id, parent_company_name, status]):
            customers = customers.filter(
                Q(name__icontains=general_search) |
                Q(registered_name__icontains=general_search) |
                Q(tin__icontains=general_search) |
                Q(city__icontains=general_search)
            )

        # Apply sorting
        sort_prefix = '-' if sort_direction == 'desc' else ''
        customers = customers.order_by(f'{sort_prefix}{sort_by}')
        
        # Pagination
        page = self.paginate_queryset(customers, request)
        if page is not None:
            serializer = CustomerSerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })

        # Fallback if pagination fails
        serializer = CustomerSerializer(customers, many=True)
        return Response({
            'success': True,
            'data': serializer.data
        })

    def post(self, request):
        serializer = CustomerCreateUpdateSerializer(data=request.data)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk):
        customer = get_object_or_404(Customer, pk=pk)
        serializer = CustomerCreateUpdateSerializer(customer, data=request.data, partial=True)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                })
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        customer = get_object_or_404(Customer, pk=pk)
        customer.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class BrokerView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        # If pk is provided, return a single broker with all related data
        if pk:
            broker = get_object_or_404(Broker, pk=pk)
            serializer = BrokerSerializer(broker)
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get search parameters for specific fields
        company_name_search = request.query_params.get('company_name', '')
        phone_number_search = request.query_params.get('phone_number', '')
        email_search = request.query_params.get('email', '')
        
        # Get other filter parameters
        payment_type = request.query_params.get('payment_type', '')
        
        # Get general search parameter
        general_search = request.query_params.get('search', '')
        
        # Get sorting parameters
        sort_by = request.query_params.get('sort_by', 'company_name')
        sort_direction = request.query_params.get('sort_direction', 'asc')
        
        # Query brokers
        brokers = Broker.objects.all()

        # Apply field-specific search filters
        if company_name_search:
            brokers = brokers.filter(company_name__icontains=company_name_search)
        
        if phone_number_search:
            brokers = brokers.filter(phone_number__icontains=phone_number_search)
        
        if email_search:
            brokers = brokers.filter(email__icontains=email_search)
        
        # Filter by payment type if provided
        if payment_type and payment_type in dict(Broker.PAYMENT_CHOICES):
            brokers = brokers.filter(payment_type=payment_type)

        # Apply general search filter if no specific filters are provided
        if general_search and not any([company_name_search, phone_number_search, email_search, payment_type]):
            brokers = brokers.filter(
                Q(company_name__icontains=general_search) |
                Q(address__icontains=general_search) |
                Q(email__icontains=general_search) |
                Q(phone_number__icontains=general_search)
            )

        # Apply sorting
        sort_prefix = '-' if sort_direction == 'desc' else ''
        brokers = brokers.order_by(f'{sort_prefix}{sort_by}')
        
        # Pagination
        page = self.paginate_queryset(brokers, request)
        if page is not None:
            serializer = BrokerSerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })

        # Fallback if pagination fails
        serializer = BrokerSerializer(brokers, many=True)
        return Response({
            'success': True,
            'data': serializer.data
        })

    def post(self, request):
        serializer = BrokerCreateUpdateSerializer(data=request.data)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk):
        broker = get_object_or_404(Broker, pk=pk)
        serializer = BrokerCreateUpdateSerializer(broker, data=request.data, partial=True)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                })
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        broker = get_object_or_404(Broker, pk=pk)
        broker.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class ForwarderView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        # If pk is provided, return a single forwarder with all related data
        if pk:
            forwarder = get_object_or_404(Forwarder, pk=pk)
            serializer = ForwarderSerializer(forwarder)
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get search parameters for specific fields
        company_name_search = request.query_params.get('company_name', '')
        phone_number_search = request.query_params.get('phone_number', '')
        email_search = request.query_params.get('email', '')
        
        # Get other filter parameters
        payment_type = request.query_params.get('payment_type', '')
        
        # Get general search parameter
        general_search = request.query_params.get('search', '')
        
        # Get sorting parameters
        sort_by = request.query_params.get('sort_by', 'company_name')
        sort_direction = request.query_params.get('sort_direction', 'asc')
        
        # Query forwarders
        forwarders = Forwarder.objects.all()

        # Apply field-specific search filters
        if company_name_search:
            forwarders = forwarders.filter(company_name__icontains=company_name_search)
        
        if phone_number_search:
            forwarders = forwarders.filter(phone_number__icontains=phone_number_search)
        
        if email_search:
            forwarders = forwarders.filter(email__icontains=email_search)
        
        # Filter by payment type if provided
        if payment_type and payment_type in dict(Forwarder.PAYMENT_CHOICES):
            forwarders = forwarders.filter(payment_type=payment_type)

        # Apply general search filter if no specific filters are provided
        if general_search and not any([company_name_search, phone_number_search, email_search, payment_type]):
            forwarders = forwarders.filter(
                Q(company_name__icontains=general_search) |
                Q(address__icontains=general_search) |
                Q(email__icontains=general_search) |
                Q(phone_number__icontains=general_search)
            )

        # Apply sorting
        sort_prefix = '-' if sort_direction == 'desc' else ''
        forwarders = forwarders.order_by(f'{sort_prefix}{sort_by}')
        
        # Pagination
        page = self.paginate_queryset(forwarders, request)
        if page is not None:
            serializer = ForwarderSerializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })

        # Fallback if pagination fails
        serializer = ForwarderSerializer(forwarders, many=True)
        return Response({
            'success': True,
            'data': serializer.data
        })

    def post(self, request):
        serializer = ForwarderCreateUpdateSerializer(data=request.data)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk):
        forwarder = get_object_or_404(Forwarder, pk=pk)
        serializer = ForwarderCreateUpdateSerializer(forwarder, data=request.data, partial=True)
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                })
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk):
        forwarder = get_object_or_404(Forwarder, pk=pk)
        forwarder.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class InventoryView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    def get(self, request, pk=None):
        # If pk is provided, return a single inventory item with all related data
        if pk:
            inventory = get_object_or_404(Inventory, pk=pk)
            serializer = InventorySerializer(inventory, context={'request': request})
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get search parameters for specific fields
        item_code_search = request.query_params.get('item_code', '')
        cip_code_search = request.query_params.get('cip_code', '')
        product_name_search = request.query_params.get('product_name', '')
        brand_search = request.query_params.get('brand', '')
        external_description_search = request.query_params.get('external_description', '')
        color_search = request.query_params.get('color', '')
        materials_search = request.query_params.get('materials', '')
        pattern_search = request.query_params.get('pattern', '')
        
        # Get other filter parameters
        status_filter = request.query_params.get('status', '')
        supplier_name = request.query_params.get('supplier', '')
        brand_name = request.query_params.get('brand', '')
        category_name = request.query_params.get('category', '')
        product_tagging_filter = request.query_params.get('product_tagging', '')
        
        # Get general search parameter
        general_search = request.query_params.get('search', '')
        
        # Get sorting parameters
        sort_by = request.query_params.get('sort_by', 'item_code')
        sort_direction = request.query_params.get('sort_direction', 'asc')
        
        # Query inventory items
        inventory_items = Inventory.objects.all()

        # Apply field-specific search filters
        if item_code_search:
            inventory_items = inventory_items.filter(item_code__icontains=item_code_search)

        if cip_code_search:
            inventory_items = inventory_items.filter(cip_code__icontains=cip_code_search)
        
        if product_name_search:
            inventory_items = inventory_items.filter(product_name__icontains=product_name_search)
        
        if brand_search:
            inventory_items = inventory_items.filter(brand__name__icontains=brand_search)
            
        if external_description_search:
            inventory_items = inventory_items.filter(external_description__icontains=external_description_search)
            
        if color_search:
            inventory_items = inventory_items.filter(color__icontains=color_search)
            
        if materials_search:
            inventory_items = inventory_items.filter(materials__icontains=materials_search)
            
        if pattern_search:
            inventory_items = inventory_items.filter(pattern__icontains=pattern_search)
        
        # Apply other filters
        if status_filter and status_filter in dict(Inventory.STATUS_CHOICES):
            inventory_items = inventory_items.filter(status=status_filter)
            
        if product_tagging_filter and product_tagging_filter in dict(Inventory.PRODUCT_TAGGING_CHOICES):
            inventory_items = inventory_items.filter(product_tagging=product_tagging_filter)
            
        if supplier_name:
            inventory_items = inventory_items.filter(supplier__name__icontains=supplier_name)
                
        if brand_name:
            inventory_items = inventory_items.filter(brand__name__icontains=brand_name)
                
        if category_name:
            inventory_items = inventory_items.filter(
                Q(category__name__icontains=category_name) |
                Q(subcategory__name__icontains=category_name) |
                Q(sub_level_category__name__icontains=category_name)
            )

        # Apply general search filter across all searchable fields
        if general_search:
            inventory_items = inventory_items.filter(
                Q(item_code__icontains=general_search) |
                Q(cip_code__icontains=general_search) |
                Q(product_name__icontains=general_search) |
                Q(brand__name__icontains=general_search) |
                Q(supplier__name__icontains=general_search) |
                Q(category__name__icontains=general_search) |
                Q(subcategory__name__icontains=general_search) |
                Q(sub_level_category__name__icontains=general_search) |
                Q(external_description__icontains=general_search) |
                Q(color__icontains=general_search) |
                Q(materials__icontains=general_search) |
                Q(pattern__icontains=general_search) |
                Q(status__icontains=general_search) |
                Q(product_tagging__icontains=general_search)
            )

        # Apply sorting
        sort_prefix = '-' if sort_direction == 'desc' else ''
        inventory_items = inventory_items.order_by(f'{sort_prefix}{sort_by}')
        
        # Pagination
        page = self.paginate_queryset(inventory_items, request)
        if page is not None:
            serializer = InventorySerializer(page, many=True, context={'request': request})
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })

        # Fallback if pagination fails
        serializer = InventorySerializer(inventory_items, many=True, context={'request': request})
        return Response({
            'success': True,
            'data': serializer.data
        })
    
    def delete(self, request, pk):
        inventory = get_object_or_404(Inventory, pk=pk)
        inventory.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class InventoryTemplateView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """
        Download an empty Excel template for inventory import
        """
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Template"
        
        # Define headers
        headers = [
            'Item Code*', 'CIP Code*', 'Product Name*', 'Status*', 'Supplier ID*', 'Brand ID*',
            'Product Tagging*', 'Audit Status*', 'Category ID*', 'Subcategory ID', 'Sub Level Category ID',
            'Unit', 'Landed Cost Price', 'Landed Cost Unit', 'Packaging Amount',
            'Packaging Units', 'Packaging Package', 'External Description',
            'Length', 'Length Unit', 'Color', 'Width', 'Width Unit',
            'Height', 'Height Unit', 'Volume', 'Volume Unit', 'Materials', 'Pattern',
            'List Price Currency', 'List Price', 'Wholesale Price', 'Remarks'
        ]
        
        # Add headers to the worksheet
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            
            # Adjust column width
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
        
        # Add a second row with example data
        example_data = [
            'ITM001', 'Example Product', 'active', '1', '1',
            'never_sold', 'False', '1', '2', '3',
            'pcs', '100.00', 'USD', '10',
            'pcs', 'Box', 'Product description here',
            '10.5', 'cm', 'Blue', '5.2', 'cm',
            '3.1', 'cm', '170.5', 'cm³', 'Wood, Metal',
            'USD', '150.00', '120.00', 'Additional notes'
        ]
        
        for col_num, value in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.font = Font(italic=True)
        
        # Create a buffer to save the workbook
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create the HttpResponse with the appropriate Excel headers
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=inventory_template.xlsx'
        
        return response

class InventoryUploadView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)
    
    def post(self, request):
        """
        Upload and process an Excel file with inventory data
        """
        if 'file' not in request.FILES:
            return Response({
                'success': False,
                'errors': {'file': 'No file was uploaded.'}
            }, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        
        # Check file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({
                'success': False,
                'errors': {'file': 'Uploaded file must be an Excel file (.xlsx or .xls).'}
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Read the Excel file
            df = pd.read_excel(excel_file)
            
            # Check if the file has data
            if df.empty:
                return Response({
                    'success': False,
                    'errors': {'file': 'The uploaded file is empty.'}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Rename columns to match model field names
            column_mapping = {
                'Item Code*': 'item_code',
                'CIP Code*': 'cip_code',
                'Product Name*': 'product_name',
                'Status*': 'status',
                'Supplier ID*': 'supplier',
                'Brand ID*': 'brand',
                'Product Tagging*': 'product_tagging',
                'Audit Status*': 'audit_status',
                'Category ID*': 'category',
                'Subcategory ID': 'subcategory',
                'Sub Level Category ID': 'sub_level_category',
                'Unit': 'unit',
                'Landed Cost Price': 'landed_cost_price',
                'Landed Cost Unit': 'landed_cost_unit',
                'Packaging Amount': 'packaging_amount',
                'Packaging Units': 'packaging_units',
                'Packaging Package': 'packaging_package',
                'External Description': 'external_description',
                'Length': 'length',
                'Length Unit': 'length_unit',
                'Color': 'color',
                'Width': 'width',
                'Width Unit': 'width_unit',
                'Height': 'height',
                'Height Unit': 'height_unit',
                'Volume': 'volume',
                'Volume Unit': 'volume_unit',
                'Materials': 'materials',
                'Pattern': 'pattern',
                'List Price Currency': 'list_price_currency',
                'List Price': 'list_price',
                'Wholesale Price': 'wholesale_price',
                'Remarks': 'remarks'
            }
            
            df = df.rename(columns=column_mapping)
            
            # Check for required columns
            required_columns = ['item_code', 'product_name', 'status', 'supplier', 'brand', 'product_tagging', 'audit_status', 'category']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return Response({
                    'success': False,
                    'errors': {'file': f'Missing required columns: {", ".join(missing_columns)}'}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Process each row
            success_count = 0
            error_rows = []

            for index, row in df.iterrows():
                row_num = index + 2  # Excel is 1-indexed + header row
                row_data = row.to_dict()
                for key, value in row_data.items():
                    if pd.isna(value):
                        row_data[key] = None

                validation_errors = {} # Initialize empty dict for errors for this row

                # --- Perform ALL validations and collect errors ---

                # Validate required fields
                for field in required_columns:
                    # Special handling for boolean audit_status where False is valid
                    if field == 'audit_status' and row_data.get(field) is not None:
                         # Allow False, None, True. Empty string or missing is caught below.
                         if isinstance(row_data.get(field), str) and row_data.get(field) == '':
                              validation_errors[field] = 'This field is required.'
                         elif row_data.get(field) is None:
                              # Check if it was originally NaN/None vs empty string
                              original_value = row[field] # Get original value from Series
                              if pd.isna(original_value): # If it was truly missing/NaN
                                   validation_errors[field] = 'This field is required.'
                              # If it was an empty string converted to None, it's caught above
                    elif not row_data.get(field) and row_data.get(field) != False: # Check for empty/None, allow False
                        validation_errors[field] = 'This field is required.'


                # Validate status (only if not already missing)
                if 'status' not in validation_errors and row_data.get('status') not in dict(Inventory.STATUS_CHOICES):
                    validation_errors['status'] = f'Status must be one of: {", ".join(dict(Inventory.STATUS_CHOICES).keys())}'

                # Validate product_tagging (only if not already missing)
                if 'product_tagging' not in validation_errors and row_data.get('product_tagging') not in dict(Inventory.PRODUCT_TAGGING_CHOICES):
                    validation_errors['product_tagging'] = f'Product Tagging must be one of: {", ".join(dict(Inventory.PRODUCT_TAGGING_CHOICES).keys())}'

                # Convert and validate audit_status (only if not already missing)
                if 'audit_status' not in validation_errors and row_data.get('audit_status') is not None:
                    audit_status_val = row_data['audit_status']
                    if isinstance(audit_status_val, str):
                        if audit_status_val.lower() in ('true', 't', 'yes', 'y', '1'):
                            row_data['audit_status'] = True
                        elif audit_status_val.lower() in ('false', 'f', 'no', 'n', '0'):
                            row_data['audit_status'] = False
                        else:
                            validation_errors['audit_status'] = 'Audit Status must be True or False.'
                    elif not isinstance(audit_status_val, bool):
                         validation_errors['audit_status'] = 'Audit Status must be True or False.'
                    # If it's already a bool (e.g., from Excel), it's fine

                # Validate foreign keys (only if not already missing)
                category_id = None # Keep track for subcategory validation
                if 'supplier' not in validation_errors:
                    try:
                        supplier_id = int(row_data.get('supplier'))
                        if not Supplier.objects.filter(id=supplier_id).exists():
                            validation_errors['supplier'] = f'Supplier with ID {supplier_id} does not exist.'
                        # No need to re-assign row_data['supplier'] here, serializer handles ID
                    except (ValueError, TypeError):
                        validation_errors['supplier'] = 'Supplier ID must be a valid number.'

                if 'brand' not in validation_errors:
                    try:
                        brand_id = int(row_data.get('brand'))
                        if not Brand.objects.filter(id=brand_id).exists():
                            validation_errors['brand'] = f'Brand with ID {brand_id} does not exist.'
                    except (ValueError, TypeError):
                        validation_errors['brand'] = 'Brand ID must be a valid number.'

                if 'category' not in validation_errors:
                    try:
                        category_id = int(row_data.get('category')) # Store for later use
                        if not Category.objects.filter(id=category_id).exists():
                            validation_errors['category'] = f'Category with ID {category_id} does not exist.'
                    except (ValueError, TypeError):
                        validation_errors['category'] = 'Category ID must be a valid number.'

                # Validate subcategory if provided and category is valid
                subcategory_id = None # Keep track for sub-level validation
                if row_data.get('subcategory') and 'category' not in validation_errors and category_id is not None:
                    try:
                        subcategory_id = int(row_data.get('subcategory'))
                        subcategory = Category.objects.filter(id=subcategory_id).first()
                        if not subcategory:
                            validation_errors['subcategory'] = f'Subcategory with ID {subcategory_id} does not exist.'
                        elif subcategory.parent_id != category_id:
                            validation_errors['subcategory'] = f'Subcategory must belong to the selected category.'
                    except (ValueError, TypeError):
                        validation_errors['subcategory'] = 'Subcategory ID must be a valid number.'
                elif row_data.get('subcategory') and ('category' in validation_errors or category_id is None):
                     # Avoid validating subcategory if category itself is invalid/missing
                     pass


                # Validate sub_level_category if provided and subcategory is valid
                if row_data.get('sub_level_category') and 'subcategory' not in validation_errors and subcategory_id is not None:
                    try:
                        sub_level_id = int(row_data.get('sub_level_category'))
                        sub_level = Category.objects.filter(id=sub_level_id).first()
                        if not sub_level:
                            validation_errors['sub_level_category'] = f'Sub Level Category with ID {sub_level_id} does not exist.'
                        elif sub_level.parent_id != subcategory_id:
                            validation_errors['sub_level_category'] = f'Sub Level Category must belong to the selected subcategory.'
                    except (ValueError, TypeError):
                        validation_errors['sub_level_category'] = 'Sub Level Category ID must be a valid number.'
                elif row_data.get('sub_level_category') and ('subcategory' in validation_errors or subcategory_id is None):
                     # Avoid validating sub-level if subcategory itself is invalid/missing
                     pass

                # Check for duplicate item_code (only if not already missing)
                if 'item_code' not in validation_errors and Inventory.objects.filter(item_code=row_data.get('item_code')).exists():
                    validation_errors['item_code'] = f'Item Code {row_data.get("item_code")} already exists.'

                # --- End of Validations ---

                # If any errors were found during manual checks, record them and skip to next row
                if validation_errors:
                    error_rows.append({
                        'row': row_num,
                        'errors': validation_errors
                    })
                    continue # Skip serializer validation/creation for this row

                # If manual checks passed, proceed with serializer validation and creation
                try:
                    # Set has_description flag
                    description_fields = [
                        'unit', 'landed_cost_price', 'landed_cost_unit', 'packaging_amount',
                        'packaging_units', 'packaging_package', 'external_description', 'length',
                        'length_unit', 'color', 'width', 'width_unit', 'height', 'height_unit',
                        'volume', 'volume_unit', 'materials', 'pattern', 'list_price_currency',
                        'list_price', 'wholesale_price', 'remarks'
                    ]
                    has_description = any(row_data.get(field) for field in description_fields)
                    row_data['has_description'] = has_description

                    # Prepare data for serializer (ensure FKs are IDs)
                    # Pandas might convert IDs to floats if there were NaNs, ensure they are ints if not None
                    for fk_field in ['supplier', 'brand', 'category', 'subcategory', 'sub_level_category']:
                         if row_data.get(fk_field) is not None:
                              try:
                                   row_data[fk_field] = int(row_data[fk_field])
                              except (ValueError, TypeError):
                                   # This case should have been caught earlier, but as a safeguard:
                                   error_rows.append({'row': row_num, 'errors': {fk_field: 'Invalid ID format.'}})
                                   validation_errors[fk_field] = 'Invalid ID format.' # Mark error to prevent saving
                                   break # Stop processing this row

                    if validation_errors: # Check again if FK conversion failed
                         continue

                    # Use serializer for creation
                    serializer = InventorySerializer(data=row_data, context={'request': request})
                    if serializer.is_valid():
                        # We already set created_by/last_modified_by in serializer context
                        serializer.save()
                        success_count += 1
                    else:
                        # Format serializer errors (should be rare if manual checks are thorough)
                        formatted_errors = {}
                        for field, errors in serializer.errors.items():
                            formatted_errors[field] = errors[0] if isinstance(errors, list) else errors
                        error_rows.append({
                            'row': row_num,
                            'errors': formatted_errors
                        })

                except Exception as e:
                    # Catch unexpected errors during save
                    error_rows.append({
                        'row': row_num,
                        'errors': {'detail': f'Unexpected error during save: {str(e)}'}
                    })

            # Return the results
            return Response({
                'success': True,
                'data': {
                    'total_rows': len(df),
                    'success_count': success_count,
                    'error_count': len(error_rows),
                    # Return all errors or limit them if needed
                    'errors': error_rows
                }
            })

        except pd.errors.EmptyDataError:
             return Response({'success': False, 'errors': {'file': 'The uploaded file is empty or could not be read.'}}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            # Catch errors during file reading, column mapping etc.
            return Response({
                'success': False,
                'errors': {'detail': f'Error processing file: {str(e)}'}
            }, status=status.HTTP_400_BAD_REQUEST)

class InventoryGeneralView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        # Extract only general fields from request data
        general_fields = [
            'item_code', 'cip_code', 'product_name', 'status', 'supplier', 'brand',
            'product_tagging', 'category', 'subcategory', 'sub_level_category'
        ]
        general_data = {k: v for k, v in request.data.items() if k in general_fields}
        
        serializer = InventorySerializer(data=general_data, context={'request': request})
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk):
        inventory = get_object_or_404(Inventory, pk=pk)
        
        # Extract only general fields from request data
        general_fields = [
            'item_code', 'cip_code', 'product_name', 'status', 'supplier', 'brand',
            'product_tagging', 'category', 'subcategory', 'sub_level_category'
        ]
        general_data = {k: v for k, v in request.data.items() if k in general_fields}
        
        serializer = InventorySerializer(inventory, data=general_data, partial=True, context={'request': request})
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                })
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)

class InventoryDescriptionView(APIView):
    permission_classes = [IsAuthenticated]
    
    def put(self, request, pk):
        inventory = get_object_or_404(Inventory, pk=pk)
        
        # Extract only description fields from request data
        description_fields = [
            'unit', 'landed_cost_price', 'landed_cost_unit', 
            'packaging_amount', 'packaging_units', 'packaging_package',
            'external_description', 'length', 'length_unit', 'color',
            'width', 'width_unit', 'height', 'height_unit',
            'volume', 'volume_unit', 'materials', 'pattern', 'photo',
            'list_price_currency', 'list_price', 'wholesale_price', 'remarks'
        ]
        
        # Handle file upload separately
        description_data = {}
        for field in description_fields:
            if field != 'photo' and field in request.data:
                description_data[field] = request.data[field]
        
        # Handle photo upload if present
        if 'photo' in request.FILES:
            description_data['photo'] = request.FILES['photo']
        
        # Set has_description to True if any description field is provided
        if any(description_data.values()):
            description_data['has_description'] = True
        
        serializer = InventorySerializer(inventory, data=description_data, partial=True, context={'request': request})
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                })
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)