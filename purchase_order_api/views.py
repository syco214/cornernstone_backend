import json
from datetime import datetime
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.core.exceptions import FieldDoesNotExist, ValidationError
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
import io
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from decimal import Decimal, InvalidOperation

from .models import PurchaseOrder, PurchaseOrderRoute, PurchaseOrderDownPayment, PackingList, PaymentDocument, InvoiceDocument, PurchaseOrderItem
from .serializers import PurchaseOrderSerializer, PurchaseOrderCreateUpdateSerializer, PurchaseOrderRouteSerializer, PurchaseOrderDownPaymentSerializer, PackingListSerializer, PaymentDocumentSerializer, InvoiceDocumentSerializer
from .po_workflows import POWorkflow
from admin_api.models import Inventory

class PurchaseOrderView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        # If pk is provided, return a single purchase order with all related data
        if pk:
            # For single object retrieval, it's good to optimize with select_related/prefetch_related
            purchase_order = get_object_or_404(
                PurchaseOrder.objects.select_related(
                    'supplier', 'created_by', 'last_modified_by', 'approved_by', 'payment_term'
                ).prefetch_related(
                    'items__inventory', 'discounts_charges'
                ), 
                pk=pk
            )
            serializer = PurchaseOrderSerializer(purchase_order) # Use the detailed serializer
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get search parameters for specific fields
        po_number_search = request.query_params.get('po_number', '')
        status_search = request.query_params.get('status', '')
        supplier_search = request.query_params.get('supplier', '') # Search by supplier name
        date_from_str = request.query_params.get('date_from', '')
        date_to_str = request.query_params.get('date_to', '')
        
        # Get general search parameter
        general_search = request.query_params.get('search', '')
        
        # Get sorting parameters
        sort_by = request.query_params.get('sort_by', '-po_date') # Default sort for POs
        sort_direction = request.query_params.get('sort_direction', 'asc') # Default direction
        
        queryset = PurchaseOrder.objects.all()

        # Apply field-specific search filters
        if po_number_search:
            queryset = queryset.filter(po_number__icontains=po_number_search)
        
        if status_search and status_search in dict(PurchaseOrder.STATUS_CHOICES):
            queryset = queryset.filter(status=status_search)
            
        if supplier_search:
            queryset = queryset.filter(supplier__name__icontains=supplier_search)
            
        if date_from_str:
            try:
                date_from_obj = datetime.strptime(date_from_str, '%Y-%m-%d').date()
                queryset = queryset.filter(po_date__gte=date_from_obj)
            except ValueError:
                pass # Ignore invalid date format
                
        if date_to_str:
            try:
                date_to_obj = datetime.strptime(date_to_str, '%Y-%m-%d').date()
                queryset = queryset.filter(po_date__lte=date_to_obj)
            except ValueError:
                pass # Ignore invalid date format

        # Apply general search filter if no specific filters are provided
        if general_search and not any([po_number_search, status_search, supplier_search, date_from_str, date_to_str]):
            queryset = queryset.filter(
                Q(po_number__icontains=general_search) |
                Q(supplier__name__icontains=general_search) |
                Q(items__inventory__item_code__icontains=general_search) | # Example related field search
                Q(items__external_description__icontains=general_search) | # Example related field search
                Q(notes__icontains=general_search) # Search in PO notes
            ).distinct()

        # Apply sorting
        sort_prefix = '-' if sort_direction == 'desc' else ''
        sort_field_cleaned = sort_by.lstrip('-') # Remove existing prefix if any
        sort_field_actual = sort_field_cleaned # Assuming sort_by uses direct model field names or valid relations
        sort_order = f'{sort_prefix}{sort_field_actual}'
        
        try:
            # Attempt to order. This will raise FieldError if sort_order is invalid.
            queryset = queryset.order_by(sort_order)
        except FieldDoesNotExist: # More specific exception for invalid field names
            # Fallback to default sort if sort_by is invalid
            queryset = queryset.order_by('-po_date') # Default sort for PO
        except Exception: # Catch any other ordering error and fallback
            queryset = queryset.order_by('-po_date')


        # Pagination
        page = self.paginate_queryset(queryset, request)

        # Prepare meta data (consistent with your QuotationView structure)
        meta_data = {
            'currency_options': [choice[0] for choice in PurchaseOrder.CURRENCY_CHOICES],
            'status_options': [choice[0] for choice in PurchaseOrder.STATUS_CHOICES],
            'supplier_type_options': [choice[0] for choice in PurchaseOrder.SUPPLIER_TYPE_CHOICES],
        }
        
        if page is not None:
            serializer = PurchaseOrderSerializer(page, many=True) # Use the detailed serializer
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    },
                    **meta_data
                }
            })

        # Fallback if pagination fails or is not used
        serializer = PurchaseOrderSerializer(queryset, many=True) # Use the detailed serializer
        return Response({
            'success': True,
            'data': serializer.data,
            'meta': meta_data
        })

    def post(self, request):
        if 'data' in request.data and isinstance(request.data['data'], str):
            try:
                payload = json.loads(request.data['data'])
            except json.JSONDecodeError:
                return Response({
                    'success': False, 
                    'errors': {'detail': 'Invalid JSON in data field.'}
                }, status=status.HTTP_400_BAD_REQUEST)
        else:
            payload = request.data

        serializer = PurchaseOrderCreateUpdateSerializer(data=payload, context={'request': request})
        if serializer.is_valid():
            try:
                purchase_order = serializer.save()
                # Return full detail using the read serializer
                return Response({
                    'success': True,
                    'data': PurchaseOrderSerializer(purchase_order).data
                }, status=status.HTTP_201_CREATED)
            except Exception as e: # Catch potential errors during save (e.g. DB constraints)
                return Response({
                    'success': False,
                    'errors': {'detail': str(e)}
                }, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, pk):
        purchase_order = get_object_or_404(PurchaseOrder, pk=pk)
        raw_data = request.data
        payload = {}

        if 'data' in raw_data and isinstance(raw_data.get('data'), (str, bytes)):
            try:
                payload = json.loads(raw_data['data'])
            except json.JSONDecodeError:
                return Response({
                    'success': False,
                    'errors': {'detail': 'Invalid JSON in data field.'}
                }, status=status.HTTP_400_BAD_REQUEST)
        elif isinstance(raw_data, dict): # If it's already a dict (e.g. application/json)
             payload = raw_data
        else:
            return Response({
                'success': False,
                'errors': {'detail': 'Invalid request payload.'}
            }, status=status.HTTP_400_BAD_REQUEST)


        serializer = PurchaseOrderCreateUpdateSerializer(
            purchase_order,
            data=payload,
            partial=True, # Allow partial updates
            context={'request': request}
        )
        if serializer.is_valid():
            try:
                updated_purchase_order = serializer.save()
                return Response({
                    'success': True,
                    'data': PurchaseOrderSerializer(updated_purchase_order).data
                })
            except Exception as e:
                return Response({
                    'success': False,
                    'errors': {'detail': str(e)}
                }, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        purchase_order = get_object_or_404(PurchaseOrder, pk=pk)
        try:
            if purchase_order.status not in ['draft', 'cancelled', 'rejected']:
                return Response({
                    'success': False,
                    'errors': {'detail': f'Cannot delete PO in {purchase_order.get_status_display()} status.'}
                }, status=status.HTTP_400_BAD_REQUEST)
            purchase_order.delete()
            return Response({
                'success': True,
                'data': None # Or a confirmation message
            }, status=status.HTTP_200_OK) # Or HTTP_204_NO_CONTENT
        except Exception as e: # Catch potential errors during delete (e.g. ProtectedError)
             return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)

class PurchaseOrderWorkflowView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    
    def post(self, request, pk, action=None):
        purchase_order = get_object_or_404(PurchaseOrder, pk=pk)
        
        # Validate the state for each action
        if action == 'submit_for_approval' and purchase_order.status != 'draft':
            return Response({
                'success': False,
                'errors': {'detail': f'Cannot submit PO in {purchase_order.get_status_display()} status. Only draft POs can be submitted.'}
            }, status=status.HTTP_400_BAD_REQUEST)
            
        elif action == 'approve_po' and purchase_order.status != 'pending_approval':
            return Response({
                'success': False,
                'errors': {'detail': f'Cannot approve PO in {purchase_order.get_status_display()} status. Only pending approval POs can be approved.'}
            }, status=status.HTTP_400_BAD_REQUEST)
            
        elif action == 'reject_po' and purchase_order.status != 'pending_approval':
            return Response({
                'success': False,
                'errors': {'detail': f'Cannot reject PO in {purchase_order.get_status_display()} status. Only pending approval POs can be rejected.'}
            }, status=status.HTTP_400_BAD_REQUEST)
            
        elif action == 'submit_dp' and purchase_order.status != 'for_dp':
            return Response({
                'success': False,
                'errors': {'detail': f'Cannot submit down payment for PO in {purchase_order.get_status_display()} status. Only POs in For Down Payment status can have down payments submitted.'}
            }, status=status.HTTP_400_BAD_REQUEST)
            
        elif action == 'approve_dp' and purchase_order.status != 'pending_dp_approval':
            return Response({
                'success': False,
                'errors': {'detail': f'Cannot approve down payment for PO in {purchase_order.get_status_display()} status. Only POs with pending down payments can be approved.'}
            }, status=status.HTTP_400_BAD_REQUEST)
            
        elif action == 'reject_dp' and purchase_order.status != 'pending_dp_approval':
            return Response({
                'success': False,
                'errors': {'detail': f'Cannot reject down payment for PO in {purchase_order.get_status_display()} status. Only POs with pending down payments can be rejected.'}
            }, status=status.HTTP_400_BAD_REQUEST)
            
        elif action == 'confirm_ready_dates' and purchase_order.status != 'confirm_ready_dates':
            return Response({
                'success': False,
                'errors': {'detail': f'Cannot confirm ready dates for PO in {purchase_order.get_status_display()} status. Only POs in Ready Date Confirmation status can proceed.'}
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Initialize workflow if needed for actions that require it
        if action in ['submit_for_approval', 'approve_po'] and not purchase_order.route_steps.exists():
            POWorkflow.initialize_workflow(purchase_order, request.user)
        
        # Handle different workflow actions
        if action == 'submit_for_approval':
            purchase_order = POWorkflow.submit_for_approval(purchase_order, request.user)
            message = "Purchase order submitted for approval successfully"
            
        elif action == 'approve_po':
            purchase_order = POWorkflow.approve_po(purchase_order, request.user)
            message = "Purchase order approved successfully"
            
        elif action == 'reject_po':
            purchase_order = POWorkflow.reject_po(purchase_order, request.user)
            message = "Purchase order rejected successfully"
            
        elif action == 'submit_dp':
            # Validate the amount_paid field
            amount_paid = request.data.get('amount_paid')
            if not amount_paid:
                return Response({
                    'success': False,
                    'errors': {'amount_paid': ['This field is required.']}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                amount_paid = float(amount_paid)
                if amount_paid <= 0:
                    return Response({
                        'success': False,
                        'errors': {'amount_paid': ['Amount must be greater than zero.']}
                    }, status=status.HTTP_400_BAD_REQUEST)
            except ValueError:
                return Response({
                    'success': False,
                    'errors': {'amount_paid': ['Invalid amount. Must be a positive number.']}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Handle down payment submission with file upload
            data = {
                'amount_paid': amount_paid,
                'remarks': request.data.get('remarks', ''),
            }
            
            if 'payment_slip' in request.FILES:
                data['payment_slip'] = request.FILES['payment_slip']
                
            purchase_order, down_payment = POWorkflow.submit_down_payment(purchase_order, data, request.user)
            
            # Create response with down payment details
            dp_serializer = PurchaseOrderDownPaymentSerializer(
                down_payment,
                context={'request': request}
            )
            
            return Response({
                'success': True,
                'message': "Down payment submitted successfully",
                'data': {
                    'purchase_order': PurchaseOrderSerializer(purchase_order).data,
                    'down_payment': dp_serializer.data
                }
            })
            
        elif action == 'approve_dp':
            purchase_order = POWorkflow.approve_down_payment(purchase_order, request.user)
            message = "Down payment approved successfully"
            
        elif action == 'reject_dp':
            purchase_order = POWorkflow.reject_down_payment(purchase_order, request.user)
            message = "Down payment rejected successfully"
            
        elif action == 'confirm_ready_dates':
            # Validate that items array is present in the request
            if not request.data.get('items'):
                return Response({
                    'success': False,
                    'errors': {'items': ['This field is required.']}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                # Call the workflow method
                purchase_order = POWorkflow.confirm_ready_dates(purchase_order, request.data, request.user)
                message = "Ready dates confirmed successfully"
            except ValidationError as e:
                return Response({
                    'success': False,
                    'errors': {'detail': str(e)}
                }, status=status.HTTP_400_BAD_REQUEST)
            
        elif action.startswith('submit_packing_list_'):
            # Extract batch number from action name
            try:
                batch_number = int(action.split('_')[-1])
            except (ValueError, IndexError):
                return Response({
                    'success': False,
                    'errors': {'detail': 'Invalid batch number in action.'}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate file uploads
            if 'document' not in request.FILES:
                return Response({
                    'success': False,
                    'errors': {'document': ['This field is required.']}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Prepare data
            data = {
                'total_weight': request.data.get('total_weight'),
                'total_packages': request.data.get('total_packages'),
                'total_volume': request.data.get('total_volume'),
                'document': request.FILES['document']
            }
            
            try:
                # Call workflow method
                packing_list = POWorkflow.submit_packing_list(
                    purchase_order, batch_number, data, request.user
                )
                
                # Return response
                serializer = PackingListSerializer(packing_list, context={'request': request})
                return Response({
                    'success': True,
                    'message': f'Packing list for batch {batch_number} submitted successfully',
                    'data': {
                        'purchase_order': PurchaseOrderSerializer(purchase_order).data,
                        'packing_list': serializer.data
                    }
                })
            except ValidationError as e:
                return Response({
                    'success': False,
                    'errors': {'detail': str(e)}
                }, status=status.HTTP_400_BAD_REQUEST)
        
        elif action.startswith('approve_import_'):
            try:
                batch_number = int(action.split('_')[-1])
            except (ValueError, IndexError):
                return Response({
                    'success': False,
                    'errors': {'detail': 'Invalid batch number in action.'}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Safer approach
            approve_value = request.data.get('approve', 'false')
            if isinstance(approve_value, bool):
                approve = approve_value
            else:
                approve = str(approve_value).lower() == 'true'
            
            try:
                # Call workflow method
                result = POWorkflow.approve_import(
                    purchase_order, batch_number, approve, request.user
                )
                
                # Return response
                status_text = 'approved' if approve else 'rejected'
                response_data = {
                    'success': True,
                    'message': f'Import for batch {batch_number} {status_text} successfully',
                    'data': {
                        'purchase_order': PurchaseOrderSerializer(purchase_order).data
                    }
                }
                
                if approve and result:
                    serializer = PackingListSerializer(result, context={'request': request})
                    response_data['data']['packing_list'] = serializer.data
                
                return Response(response_data)
            except ValidationError as e:
                return Response({
                    'success': False,
                    'errors': {'detail': str(e)}
                }, status=status.HTTP_400_BAD_REQUEST)
        
        elif action.startswith('submit_payment_'):
            try:
                batch_number = int(action.split('_')[-1])
            except (ValueError, IndexError):
                return Response({
                    'success': False,
                    'errors': {'detail': 'Invalid batch number in action.'}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate file uploads
            if 'document' not in request.FILES:
                return Response({
                    'success': False,
                    'errors': {'document': ['This field is required.']}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Prepare data
            data = {'document': request.FILES['document']}
            
            try:
                # Call workflow method
                payment = POWorkflow.submit_payment(
                    purchase_order, batch_number, data, request.user
                )
                
                # Return response
                serializer = PaymentDocumentSerializer(payment, context={'request': request})
                return Response({
                    'success': True,
                    'message': f'Payment for batch {batch_number} submitted successfully',
                    'data': {
                        'purchase_order': PurchaseOrderSerializer(purchase_order).data,
                        'payment': serializer.data
                    }
                })
            except ValidationError as e:
                return Response({
                    'success': False,
                    'errors': {'detail': str(e)}
                }, status=status.HTTP_400_BAD_REQUEST)
        
        elif action.startswith('submit_invoice_'):
            try:
                batch_number = int(action.split('_')[-1])
            except (ValueError, IndexError):
                return Response({
                    'success': False,
                    'errors': {'detail': 'Invalid batch number in action.'}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate file uploads
            if 'document' not in request.FILES:
                return Response({
                    'success': False,
                    'errors': {'document': ['This field is required.']}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Prepare data
            data = {'document': request.FILES['document']}
            
            try:
                # Call workflow method
                invoice = POWorkflow.submit_invoice(
                    purchase_order, batch_number, data, request.user
                )
                
                # Return response
                serializer = InvoiceDocumentSerializer(invoice, context={'request': request})
                return Response({
                    'success': True,
                    'message': f'Invoice for batch {batch_number} submitted successfully',
                    'data': {
                        'purchase_order': PurchaseOrderSerializer(purchase_order).data,
                        'invoice': serializer.data
                    }
                })
            except ValidationError as e:
                return Response({
                    'success': False,
                    'errors': {'detail': str(e)}
                }, status=status.HTTP_400_BAD_REQUEST)
        
        else:
            return Response({
                'success': False,
                'errors': {'detail': f'Unknown action: {action}'}
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Return updated PO
        serializer = PurchaseOrderSerializer(purchase_order)
        return Response({
            'success': True,
            'message': message,
            'data': serializer.data
        })

class PurchaseOrderRouteView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, po_id):
        """Get route steps for a specific purchase order"""
        purchase_order = get_object_or_404(PurchaseOrder, pk=po_id)
        
        # Get the route steps for this purchase order
        route_steps = PurchaseOrderRoute.objects.filter(purchase_order=purchase_order)
        
        # Create steps if they don't exist
        if not route_steps.exists():
            # Initialize the workflow
            POWorkflow.initialize_workflow(purchase_order)
            # Refresh the queryset
            route_steps = PurchaseOrderRoute.objects.filter(purchase_order=purchase_order)
        
        # Serialize the data
        serializer = PurchaseOrderRouteSerializer(route_steps, many=True)
        
        return Response({
            'success': True,
            'data': serializer.data
        })

class PurchaseOrderItemsTemplateView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, pk=None):
        """Download an Excel template for bulk uploading purchase order items"""
        # Verify the purchase order exists
        purchase_order = get_object_or_404(PurchaseOrder, pk=pk)
        
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Items Template"
        
        # Add headers
        headers = ['item_code', 'quantity']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
        
        # Add example row
        ws['A2'] = 'ABC123'
        ws['B2'] = 1
        
        # Save to a BytesIO object
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="purchase_order_{purchase_order.po_number}_items_template.xlsx"'
        
        return response

class PurchaseOrderItemsUploadView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request, pk=None):
        """Upload an Excel file with item codes and quantities to add to the purchase order"""
        try:
            purchase_order = get_object_or_404(PurchaseOrder, pk=pk)
            
            if 'file' not in request.FILES:
                return Response(
                    {'success': False, 'errors': 'No file uploaded'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            file = request.FILES['file']
            
            # Check file extension
            if not (file.name.endswith('.xlsx') or file.name.endswith('.xls')):
                return Response(
                    {'success': False, 'errors': 'File must be an Excel file (.xlsx or .xls)'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Process the Excel file
            try:
                # Load the workbook
                wb = load_workbook(filename=io.BytesIO(file.read()))
                ws = wb.active
                
                # Get headers from the first row
                headers = [cell.value for cell in ws[1]]
                
                # Validate required columns
                required_columns = ['item_code', 'quantity']
                for col in required_columns:
                    if col not in headers:
                        return Response(
                            {'success': False, 'errors': f'Missing required column: {col}'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                
                # Get column indices
                item_code_idx = headers.index('item_code')
                quantity_idx = headers.index('quantity')
                
                # Process each row
                results = {
                    'success': True,
                    'added': 0,
                    'errors': [],
                    'total_rows': ws.max_row - 1  # Subtract header row
                }
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    # Validate item_code
                    item_code = str(row[item_code_idx] or '').strip()
                    if not item_code:
                        results['errors'].append(f'Line {row_idx}: Item code is empty')
                        continue
                    
                    # Validate quantity
                    try:
                        quantity_value = row[quantity_idx]
                        if quantity_value is None:
                            results['errors'].append(f'Line {row_idx}: Quantity is empty')
                            continue
                            
                        # Convert to Decimal instead of float
                        quantity = Decimal(str(quantity_value))
                        if quantity <= 0:
                            results['errors'].append(f'Line {row_idx}: Quantity must be a positive number')
                            continue
                    except (ValueError, TypeError, InvalidOperation):
                        results['errors'].append(f'Line {row_idx}: Invalid quantity format')
                        continue
                    
                    # Find inventory item
                    try:
                        inventory = Inventory.objects.get(item_code=item_code)
                    except Inventory.DoesNotExist:
                        results['errors'].append(f'Line {row_idx}: Item code "{item_code}" not found')
                        continue
                    
                    # Create purchase order item
                    try:
                        # Check if item already exists in this purchase order
                        existing_item = PurchaseOrderItem.objects.filter(
                            purchase_order=purchase_order, 
                            inventory=inventory
                        ).first()
                        
                        if existing_item:
                            # Update quantity if item already exists
                            existing_item.quantity = quantity
                            existing_item.save()
                        else:
                            # Create new item
                            PurchaseOrderItem.objects.create(
                                purchase_order=purchase_order,
                                inventory=inventory,
                                quantity=quantity,
                                list_price=Decimal(str(inventory.wholesale_price)) if inventory.wholesale_price else Decimal('0.00'),
                                unit=inventory.unit,
                                external_description=inventory.external_description
                            )
                        
                        results['added'] += 1
                    except Exception as e:
                        results['errors'].append(f'Line {row_idx}: Failed to add item - {str(e)}')
                
                # Update purchase order totals
                purchase_order.update_totals()
                
                return Response(results)
                
            except Exception as e:
                import traceback
                print(traceback.format_exc())
                return Response(
                    {'success': False, 'errors': f'Error processing Excel file: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return Response(
                {'success': False, 'errors': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )