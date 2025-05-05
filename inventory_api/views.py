import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.http import HttpResponse

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.pagination import PageNumberPagination

# Imports from other apps/modules
from admin_api.models import Inventory, Supplier, Brand, Category # Reuse models
from admin_api.serializers import SupplierSerializer # Import SupplierSerializer

# Local imports
from .serializers import (
    InventorySerializer,
)

class InventoryView(APIView, PageNumberPagination):
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)
    page_size_query_param = 'page_size' # Define the query param for page size
    max_page_size = 100 # Define the maximum page size allowed

    def get_page_size(self, request):
        """
        Allow client to set page size via query param, restricted to specific values.
        """
        page_size = request.query_params.get(self.page_size_query_param)
        default_page_size = self.page_size # Get default from settings via PageNumberPagination

        if page_size:
            try:
                page_size = int(page_size)
                # Validate against allowed page sizes
                if page_size in [10, 25, 50, 100]:
                    return min(page_size, self.max_page_size)
                else:
                    # If invalid value provided, fall back to default
                    return default_page_size
            except ValueError:
                # If non-integer value provided, fall back to default
                return default_page_size
        
        # If page_size param not provided, use default
        return default_page_size

    def get(self, request, pk=None):
        # If pk is provided, return a single inventory item with all related data
        if pk:
            inventory = get_object_or_404(Inventory, pk=pk)
            serializer = InventorySerializer(inventory, context={'request': request})
            return Response({
                'success': True,
                'data': serializer.data
            })
        
        # Get search parameters for specific fields
        item_code_search = request.query_params.get('item_code', '')
        cip_code_search = request.query_params.get('cip_code', '')
        product_name_search = request.query_params.get('product_name', '')
        brand_search = request.query_params.get('brand', '')
        
        # Get other filter parameters
        status_filter = request.query_params.get('status', '')
        supplier_id = request.query_params.get('supplier_id', '')
        brand_id = request.query_params.get('brand_id', '')
        category_id = request.query_params.get('category_id', '')
        
        # Get general search parameter
        general_search = request.query_params.get('search', '')
        
        # Get sorting parameters
        sort_by = request.query_params.get('sort_by', 'item_code')
        sort_direction = request.query_params.get('sort_direction', 'asc')
        
        # Query inventory items
        inventory_items = Inventory.objects.all()

        # Apply field-specific search filters
        if item_code_search:
            inventory_items = inventory_items.filter(item_code__icontains=item_code_search)

        if cip_code_search:
            inventory_items = inventory_items.filter(cip_code__icontains=cip_code_search)
        
        if product_name_search:
            inventory_items = inventory_items.filter(product_name__icontains=product_name_search)
        
        if brand_search:
            inventory_items = inventory_items.filter(brand__name__icontains=brand_search)
        
        # Apply other filters
        if status_filter and status_filter in dict(Inventory.STATUS_CHOICES):
            inventory_items = inventory_items.filter(status=status_filter)
            
        if supplier_id:
            try:
                supplier_id = int(supplier_id)
                inventory_items = inventory_items.filter(supplier_id=supplier_id)
            except ValueError:
                pass
                
        if brand_id:
            try:
                brand_id = int(brand_id)
                inventory_items = inventory_items.filter(brand_id=brand_id)
            except ValueError:
                pass
                
        if category_id:
            try:
                category_id = int(category_id)
                inventory_items = inventory_items.filter(
                    Q(category_id=category_id) |
                    Q(subcategory_id=category_id) |
                    Q(sub_level_category_id=category_id)
                )
            except ValueError:
                pass

        # Apply general search filter if no specific filters are provided
        if general_search and not any([item_code_search, cip_code_search, product_name_search, brand_search, 
                                      status_filter, supplier_id, brand_id, category_id]):
            inventory_items = inventory_items.filter(
                Q(item_code__icontains=general_search) |
                Q(cip_code__icontains=general_search) |
                Q(product_name__icontains=general_search) |
                Q(product_tagging__icontains=general_search)
            )

        # Apply sorting
        sort_prefix = '-' if sort_direction == 'desc' else ''
        inventory_items = inventory_items.order_by(f'{sort_prefix}{sort_by}')
        
        # Pagination
        page = self.paginate_queryset(inventory_items, request)
        if page is not None:
            serializer = InventorySerializer(page, many=True, context={'request': request})
            paginated_response = self.get_paginated_response(serializer.data)
            
            return Response({
                'success': True,
                'data': paginated_response.data['results'],
                'meta': {
                    'pagination': {
                        'count': paginated_response.data['count'],
                        'next': paginated_response.data['next'],
                        'previous': paginated_response.data['previous'],
                    }
                }
            })

        # Fallback if pagination fails
        serializer = InventorySerializer(inventory_items, many=True, context={'request': request})
        return Response({
            'success': True,
            'data': serializer.data
        })
    
    def delete(self, request, pk):
        inventory = get_object_or_404(Inventory, pk=pk)
        inventory.delete()
        return Response({
            'success': True,
            'data': None
        }, status=status.HTTP_200_OK)

class InventoryTemplateView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """
        Download an empty Excel template for inventory import
        """
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Template"
        
        # Define headers
        headers = [
            'Item Code*', 'CIP Code*', 'Product Name*', 'Status*', 'Supplier ID*', 'Brand ID*',
            'Product Tagging*', 'Audit Status*', 'Category ID*', 'Subcategory ID', 'Sub Level Category ID',
            'Unit', 'Landed Cost Price', 'Landed Cost Unit', 'Packaging Amount',
            'Packaging Units', 'Packaging Package', 'External Description',
            'Length', 'Length Unit', 'Color', 'Width', 'Width Unit',
            'Height', 'Height Unit', 'Volume', 'Volume Unit', 'Materials',
            'List Price Currency', 'List Price', 'Wholesale Price', 'Remarks'
        ]
        
        # Add headers to the worksheet
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            
            # Adjust column width
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
        
        # Add a second row with example data
        example_data = [
            'ITM001', 'Example Product', 'active', '1', '1',
            'never_sold', 'False', '1', '2', '3',
            'pcs', '100.00', 'USD', '10',
            'pcs', 'Box', 'Product description here',
            '10.5', 'cm', 'Blue', '5.2', 'cm',
            '3.1', 'cm', '170.5', 'cm³', 'Wood, Metal',
            'USD', '150.00', '120.00', 'Additional notes'
        ]
        
        for col_num, value in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.font = Font(italic=True)
        
        # Create a buffer to save the workbook
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create the HttpResponse with the appropriate Excel headers
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=inventory_template.xlsx'
        
        return response

class InventoryUploadView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)
    
    def post(self, request):
        """
        Upload and process an Excel file with inventory data
        """
        if 'file' not in request.FILES:
            return Response({
                'success': False,
                'errors': {'file': 'No file was uploaded.'}
            }, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        
        # Check file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({
                'success': False,
                'errors': {'file': 'Uploaded file must be an Excel file (.xlsx or .xls).'}
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Read the Excel file
            df = pd.read_excel(excel_file)
            
            # Check if the file has data
            if df.empty:
                return Response({
                    'success': False,
                    'errors': {'file': 'The uploaded file is empty.'}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Rename columns to match model field names
            column_mapping = {
                'Item Code*': 'item_code',
                'CIP Code*': 'cip_code',
                'Product Name*': 'product_name',
                'Status*': 'status',
                'Supplier ID*': 'supplier',
                'Brand ID*': 'brand',
                'Product Tagging*': 'product_tagging',
                'Audit Status*': 'audit_status',
                'Category ID*': 'category',
                'Subcategory ID': 'subcategory',
                'Sub Level Category ID': 'sub_level_category',
                'Unit': 'unit',
                'Landed Cost Price': 'landed_cost_price',
                'Landed Cost Unit': 'landed_cost_unit',
                'Packaging Amount': 'packaging_amount',
                'Packaging Units': 'packaging_units',
                'Packaging Package': 'packaging_package',
                'External Description': 'external_description',
                'Length': 'length',
                'Length Unit': 'length_unit',
                'Color': 'color',
                'Width': 'width',
                'Width Unit': 'width_unit',
                'Height': 'height',
                'Height Unit': 'height_unit',
                'Volume': 'volume',
                'Volume Unit': 'volume_unit',
                'Materials': 'materials',
                'List Price Currency': 'list_price_currency',
                'List Price': 'list_price',
                'Wholesale Price': 'wholesale_price',
                'Remarks': 'remarks'
            }
            
            df = df.rename(columns=column_mapping)
            
            # Check for required columns
            required_columns = ['item_code', 'product_name', 'status', 'supplier', 'brand', 'product_tagging', 'audit_status', 'category']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return Response({
                    'success': False,
                    'errors': {'file': f'Missing required columns: {", ".join(missing_columns)}'}
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Process each row
            success_count = 0
            error_rows = []

            for index, row in df.iterrows():
                row_num = index + 2  # Excel is 1-indexed + header row
                row_data = row.to_dict()
                for key, value in row_data.items():
                    if pd.isna(value):
                        row_data[key] = None

                validation_errors = {} # Initialize empty dict for errors for this row

                # --- Perform ALL validations and collect errors ---

                # Validate required fields
                for field in required_columns:
                    # Special handling for boolean audit_status where False is valid
                    if field == 'audit_status' and row_data.get(field) is not None:
                         # Allow False, None, True. Empty string or missing is caught below.
                         if isinstance(row_data.get(field), str) and row_data.get(field) == '':
                              validation_errors[field] = 'This field is required.'
                         elif row_data.get(field) is None:
                              # Check if it was originally NaN/None vs empty string
                              original_value = row[field] # Get original value from Series
                              if pd.isna(original_value): # If it was truly missing/NaN
                                   validation_errors[field] = 'This field is required.'
                              # If it was an empty string converted to None, it's caught above
                    elif not row_data.get(field) and row_data.get(field) != False: # Check for empty/None, allow False
                        validation_errors[field] = 'This field is required.'


                # Validate status (only if not already missing)
                if 'status' not in validation_errors and row_data.get('status') not in dict(Inventory.STATUS_CHOICES):
                    validation_errors['status'] = f'Status must be one of: {", ".join(dict(Inventory.STATUS_CHOICES).keys())}'

                # Validate product_tagging (only if not already missing)
                if 'product_tagging' not in validation_errors and row_data.get('product_tagging') not in dict(Inventory.PRODUCT_TAGGING_CHOICES):
                    validation_errors['product_tagging'] = f'Product Tagging must be one of: {", ".join(dict(Inventory.PRODUCT_TAGGING_CHOICES).keys())}'

                # Convert and validate audit_status (only if not already missing)
                if 'audit_status' not in validation_errors and row_data.get('audit_status') is not None:
                    audit_status_val = row_data['audit_status']
                    if isinstance(audit_status_val, str):
                        if audit_status_val.lower() in ('true', 't', 'yes', 'y', '1'):
                            row_data['audit_status'] = True
                        elif audit_status_val.lower() in ('false', 'f', 'no', 'n', '0'):
                            row_data['audit_status'] = False
                        else:
                            validation_errors['audit_status'] = 'Audit Status must be True or False.'
                    elif not isinstance(audit_status_val, bool):
                         validation_errors['audit_status'] = 'Audit Status must be True or False.'
                    # If it's already a bool (e.g., from Excel), it's fine

                # Validate foreign keys (only if not already missing)
                category_id = None # Keep track for subcategory validation
                if 'supplier' not in validation_errors:
                    try:
                        supplier_id = int(row_data.get('supplier'))
                        if not Supplier.objects.filter(id=supplier_id).exists():
                            validation_errors['supplier'] = f'Supplier with ID {supplier_id} does not exist.'
                        # No need to re-assign row_data['supplier'] here, serializer handles ID
                    except (ValueError, TypeError):
                        validation_errors['supplier'] = 'Supplier ID must be a valid number.'

                if 'brand' not in validation_errors:
                    try:
                        brand_id = int(row_data.get('brand'))
                        if not Brand.objects.filter(id=brand_id).exists():
                            validation_errors['brand'] = f'Brand with ID {brand_id} does not exist.'
                    except (ValueError, TypeError):
                        validation_errors['brand'] = 'Brand ID must be a valid number.'

                if 'category' not in validation_errors:
                    try:
                        category_id = int(row_data.get('category')) # Store for later use
                        if not Category.objects.filter(id=category_id).exists():
                            validation_errors['category'] = f'Category with ID {category_id} does not exist.'
                    except (ValueError, TypeError):
                        validation_errors['category'] = 'Category ID must be a valid number.'

                # Validate subcategory if provided and category is valid
                subcategory_id = None # Keep track for sub-level validation
                if row_data.get('subcategory') and 'category' not in validation_errors and category_id is not None:
                    try:
                        subcategory_id = int(row_data.get('subcategory'))
                        subcategory = Category.objects.filter(id=subcategory_id).first()
                        if not subcategory:
                            validation_errors['subcategory'] = f'Subcategory with ID {subcategory_id} does not exist.'
                        elif subcategory.parent_id != category_id:
                            validation_errors['subcategory'] = f'Subcategory must belong to the selected category (ID: {category_id}).'
                    except (ValueError, TypeError):
                        validation_errors['subcategory'] = 'Subcategory ID must be a valid number.'
                elif row_data.get('subcategory') and ('category' in validation_errors or category_id is None):
                     # Avoid validating subcategory if category itself is invalid/missing
                     pass


                # Validate sub_level_category if provided and subcategory is valid
                if row_data.get('sub_level_category') and 'subcategory' not in validation_errors and subcategory_id is not None:
                    try:
                        sub_level_id = int(row_data.get('sub_level_category'))
                        sub_level = Category.objects.filter(id=sub_level_id).first()
                        if not sub_level:
                            validation_errors['sub_level_category'] = f'Sub Level Category with ID {sub_level_id} does not exist.'
                        elif sub_level.parent_id != subcategory_id:
                            validation_errors['sub_level_category'] = f'Sub Level Category must belong to the selected subcategory (ID: {subcategory_id}).'
                    except (ValueError, TypeError):
                        validation_errors['sub_level_category'] = 'Sub Level Category ID must be a valid number.'
                elif row_data.get('sub_level_category') and ('subcategory' in validation_errors or subcategory_id is None):
                     # Avoid validating sub-level if subcategory itself is invalid/missing
                     pass

                # Check for duplicate item_code (only if not already missing)
                if 'item_code' not in validation_errors and Inventory.objects.filter(item_code=row_data.get('item_code')).exists():
                    validation_errors['item_code'] = f'Item Code {row_data.get("item_code")} already exists.'

                # --- End of Validations ---

                # If any errors were found during manual checks, record them and skip to next row
                if validation_errors:
                    error_rows.append({
                        'row': row_num,
                        'errors': validation_errors
                    })
                    continue # Skip serializer validation/creation for this row

                # If manual checks passed, proceed with serializer validation and creation
                try:
                    # Set has_description flag
                    description_fields = [
                        'unit', 'landed_cost_price', 'landed_cost_unit', 'packaging_amount',
                        'packaging_units', 'packaging_package', 'external_description', 'length',
                        'length_unit', 'color', 'width', 'width_unit', 'height', 'height_unit',
                        'volume', 'volume_unit', 'materials', 'list_price_currency',
                        'list_price', 'wholesale_price', 'remarks'
                    ]
                    has_description = any(row_data.get(field) for field in description_fields)
                    row_data['has_description'] = has_description

                    # Prepare data for serializer (ensure FKs are IDs)
                    # Pandas might convert IDs to floats if there were NaNs, ensure they are ints if not None
                    for fk_field in ['supplier', 'brand', 'category', 'subcategory', 'sub_level_category']:
                         if row_data.get(fk_field) is not None:
                              try:
                                   row_data[fk_field] = int(row_data[fk_field])
                              except (ValueError, TypeError):
                                   # This case should have been caught earlier, but as a safeguard:
                                   error_rows.append({'row': row_num, 'errors': {fk_field: 'Invalid ID format.'}})
                                   validation_errors[fk_field] = 'Invalid ID format.' # Mark error to prevent saving
                                   break # Stop processing this row

                    if validation_errors: # Check again if FK conversion failed
                         continue

                    # Use serializer for creation
                    serializer = InventorySerializer(data=row_data, context={'request': request})
                    if serializer.is_valid():
                        # We already set created_by/last_modified_by in serializer context
                        serializer.save()
                        success_count += 1
                    else:
                        # Format serializer errors (should be rare if manual checks are thorough)
                        formatted_errors = {}
                        for field, errors in serializer.errors.items():
                            formatted_errors[field] = errors[0] if isinstance(errors, list) else errors
                        error_rows.append({
                            'row': row_num,
                            'errors': formatted_errors
                        })

                except Exception as e:
                    # Catch unexpected errors during save
                    error_rows.append({
                        'row': row_num,
                        'errors': {'detail': f'Unexpected error during save: {str(e)}'}
                    })

            # Return the results
            return Response({
                'success': True,
                'data': {
                    'total_rows': len(df),
                    'success_count': success_count,
                    'error_count': len(error_rows),
                    # Limit errors in response if needed, e.g., errors=error_rows[:50]
                    'errors': error_rows
                }
            })

        except pd.errors.EmptyDataError:
             return Response({'success': False, 'errors': {'file': 'The uploaded file is empty or could not be read.'}}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            # Catch errors during file reading, column mapping etc.
            return Response({
                'success': False,
                'errors': {'detail': f'Error processing file: {str(e)}'}
            }, status=status.HTTP_400_BAD_REQUEST)

class InventoryGeneralView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        # Extract only general fields from request data
        general_fields = [
            'item_code', 'cip_code', 'product_name', 'status', 'supplier', 'brand',
            'product_tagging', 'category', 'subcategory', 'sub_level_category'
        ]
        general_data = {k: v for k, v in request.data.items() if k in general_fields}
        
        serializer = InventorySerializer(data=general_data, context={'request': request})
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk):
        inventory = get_object_or_404(Inventory, pk=pk)
        
        # Extract only general fields from request data
        general_fields = [
            'item_code', 'cip_code', 'product_name', 'status', 'supplier', 'brand',
            'product_tagging', 'category', 'subcategory', 'sub_level_category'
        ]
        general_data = {k: v for k, v in request.data.items() if k in general_fields}
        
        serializer = InventorySerializer(inventory, data=general_data, partial=True, context={'request': request})
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                })
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)

class InventoryDescriptionView(APIView):
    permission_classes = [IsAuthenticated]
    
    def put(self, request, pk):
        inventory = get_object_or_404(Inventory, pk=pk)
        
        # Extract only description fields from request data
        description_fields = [
            'unit', 'landed_cost_price', 'landed_cost_unit', 
            'packaging_amount', 'packaging_units', 'packaging_package',
            'external_description', 'length', 'length_unit', 'color',
            'width', 'width_unit', 'height', 'height_unit',
            'volume', 'volume_unit', 'materials', 'photo',
            'list_price_currency', 'list_price', 'wholesale_price', 'remarks'
        ]
        
        # Handle file upload separately
        description_data = {}
        for field in description_fields:
            if field != 'photo' and field in request.data:
                description_data[field] = request.data[field]
        
        # Handle photo upload if present
        if 'photo' in request.FILES:
            description_data['photo'] = request.FILES['photo']
        
        # Set has_description to True if any description field is provided
        if any(description_data.values()):
            description_data['has_description'] = True
        
        serializer = InventorySerializer(inventory, data=description_data, partial=True, context={'request': request})
        try:
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'data': serializer.data
                })
            else:
                # Format validation errors
                error_messages = {}
                for field, errors in serializer.errors.items():
                    error_messages[field] = errors[0] if isinstance(errors, list) else errors
                return Response({
                    'success': False,
                    'errors': error_messages
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'errors': {'detail': str(e)}
            }, status=status.HTTP_400_BAD_REQUEST)

# Add new view for listing suppliers
class SupplierListView(APIView):
    """
    Provides a list of suppliers with their IDs and names.
    Used for frontend validation or selection.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        suppliers = Supplier.objects.all().order_by('name')
        # Use the existing SupplierSerializer if available and suitable, 
        # otherwise create a simple one or return dicts directly.
        # Assuming SupplierSerializer exists and includes 'id' and 'name'.
        serializer = SupplierSerializer(suppliers, many=True, fields=('id', 'name')) 
        return Response({
            'success': True,
            'data': serializer.data,
        })