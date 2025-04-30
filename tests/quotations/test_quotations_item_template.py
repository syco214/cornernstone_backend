import io
import tempfile
from django.test import TestCase
from django.urls import reverse
from rest_framework.test import APIClient
from rest_framework import status
from django.contrib.auth import get_user_model
from django.utils import timezone
from admin_api.models import Customer, Inventory, Supplier, Brand, Category
from quotations_api.models import Quotation, QuotationItem
from openpyxl import load_workbook, Workbook
from decimal import Decimal
import datetime

User = get_user_model()

class QuotationItemTemplateTests(TestCase):
    """Tests for quotation item template download and upload."""
    
    def setUp(self):
        """Set up test data."""
        # Create test user
        self.user = User.objects.create_user(
            username='testuser',
            password='testpassword123',
            is_staff=True
        )
        
        # Create test customer
        self.customer = Customer.objects.create(
            name='Test Customer',
            registered_name='Test Registered',
            phone_number='123-456-7890',
            company_address='123 Test St',
            city='Test City'
        )
        
        # Create test quotation with required fields
        today = timezone.now().date()
        expiry_date = today + datetime.timedelta(days=30)
        
        self.quotation = Quotation.objects.create(
            customer=self.customer,
            quote_number='QT-2023-001',
            status='draft',
            created_by=self.user,
            date=today,
            expiry_date=expiry_date,
            total_amount=Decimal('0.00'),
            currency='USD'
        )
        
        # Create test category hierarchy
        self.category = Category.objects.create(name='Electronics')
        self.subcategory = Category.objects.create(name='Computers', parent=self.category)
        
        # Create test supplier and brand
        self.supplier = Supplier.objects.create(
            name='Test Supplier',
            supplier_type='local',
            currency='USD',
            phone_number='123-456-7890',
            email='supplier@example.com'
        )
        
        self.brand = Brand.objects.create(
            name='Test Brand',
            made_in='USA'
        )
        
        # Create test inventory items
        self.inventory1 = Inventory.objects.create(
            item_code='ITEM001',
            cip_code='CIP001',
            product_name='Test Product 1',
            status='active',
            supplier=self.supplier,
            brand=self.brand,
            category=self.category,
            subcategory=self.subcategory,
            unit='pcs',
            wholesale_price=Decimal('100.00'),
            external_description='Test description 1'
        )
        
        self.inventory2 = Inventory.objects.create(
            item_code='ITEM002',
            cip_code='CIP002',
            product_name='Test Product 2',
            status='active',
            supplier=self.supplier,
            brand=self.brand,
            category=self.category,
            subcategory=self.subcategory,
            unit='pcs',
            wholesale_price=Decimal('200.00'),
            external_description='Test description 2'
        )
        
        # Set up API client
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)
        
        # URLs
        self.download_template_url = reverse('quotation-download-template', args=[self.quotation.id])
        self.upload_items_url = reverse('quotation-upload-items', args=[self.quotation.id])
    
    def test_download_template(self):
        """Test downloading the quotation items template."""
        response = self.client.get(self.download_template_url)
        
        # Check response status and content type
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn(f'quotation_{self.quotation.quote_number}_items_template.xlsx', 
                     response['Content-Disposition'])
        
        # Load the Excel file and check its contents
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check headers
        self.assertEqual(ws['A1'].value, 'item_code')
        self.assertEqual(ws['B1'].value, 'quantity')
        
        # Check example row
        self.assertEqual(ws['A2'].value, 'ABC123')
        self.assertEqual(ws['B2'].value, 1)
    
    def test_upload_items_success(self):
        """Test successfully uploading items to a quotation."""
        # Create a test Excel file
        wb = Workbook()
        ws = wb.active
        
        # Add headers
        ws['A1'] = 'item_code'
        ws['B1'] = 'quantity'
        
        # Add data rows
        ws['A2'] = 'ITEM001'
        ws['B2'] = 5
        
        # Save to a temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        # Upload the file
        with open(tmp_path, 'rb') as f:
            response = self.client.post(
                self.upload_items_url,
                {'file': f},
                format='multipart'
            )
        
        # Check response
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['added'], 1)
        
        # Verify item was added to quotation
        self.assertEqual(self.quotation.items.count(), 1)
        item = self.quotation.items.first()
        self.assertEqual(item.inventory.item_code, 'ITEM001')
        self.assertEqual(item.quantity, 5)
    
    def test_upload_items_update_existing(self):
        """Test updating existing items when uploading."""
        # Create an existing quotation item
        QuotationItem.objects.create(
            quotation=self.quotation,
            inventory=self.inventory1,
            quantity=2,
            wholesale_price=self.inventory1.wholesale_price,
            unit=self.inventory1.unit,
            external_description=self.inventory1.external_description
        )
        
        # Create a test Excel file with updated quantity
        wb = Workbook()
        ws = wb.active
        
        # Add headers
        ws['A1'] = 'item_code'
        ws['B1'] = 'quantity'
        
        # Add data row with updated quantity
        ws['A2'] = 'ITEM001'
        ws['B2'] = 10
        
        # Save to a temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        # Upload the file
        with open(tmp_path, 'rb') as f:
            response = self.client.post(
                self.upload_items_url,
                {'file': f},
                format='multipart'
            )
        
        # Check response
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['added'], 1)
        
        # Verify item was updated
        self.assertEqual(self.quotation.items.count(), 1)
        item = self.quotation.items.first()
        self.assertEqual(item.quantity, 10)
    
    def test_upload_items_invalid_item_code(self):
        """Test uploading with an invalid item code."""
        # Create a test Excel file
        wb = Workbook()
        ws = wb.active
        
        # Add headers
        ws['A1'] = 'item_code'
        ws['B1'] = 'quantity'
        
        # Add data row with invalid item code
        ws['A2'] = 'NONEXISTENT'
        ws['B2'] = 5
        
        # Save to a temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        # Upload the file
        with open(tmp_path, 'rb') as f:
            response = self.client.post(
                self.upload_items_url,
                {'file': f},
                format='multipart'
            )
        
        # Check response
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['added'], 0)
        self.assertEqual(len(response.data['errors']), 1)
        self.assertIn('not found', response.data['errors'][0])
        
        # Verify no items were added
        self.assertEqual(self.quotation.items.count(), 0)
    
    def test_upload_no_file(self):
        """Test uploading with no file."""
        response = self.client.post(self.upload_items_url, {}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertEqual(response.data['errors'], 'No file uploaded')
    
    def test_upload_invalid_file_type(self):
        """Test uploading an invalid file type."""
        # Create a text file
        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as tmp:
            tmp.write(b'This is not an Excel file')
            tmp_path = tmp.name
        
        with open(tmp_path, 'rb') as f:
            response = self.client.post(
                self.upload_items_url,
                {'file': f},
                format='multipart'
            )
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('must be an Excel file', response.data['errors'])
    
    def test_unauthorized_access(self):
        """Test that unauthenticated users cannot access the endpoints."""
        # Create a client without authentication
        client = APIClient()
        
        # Try to download template
        response = client.get(self.download_template_url)
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
        
        # Try to upload items
        response = client.post(self.upload_items_url, {}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)