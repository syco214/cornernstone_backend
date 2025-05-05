import io
from decimal import Decimal
from django.test import TestCase
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APIClient
from rest_framework import status
import openpyxl

from admin_api.models import Inventory, Supplier, Brand, Category

User = get_user_model()

class InventoryTests(TestCase):
    """Tests for the Inventory API endpoints."""
    
    def setUp(self):
        """Set up test data."""
        # Create test user
        self.user = User.objects.create_user(
            username='testuser',
            password='testpassword',
            first_name='Test',
            last_name='User'
        )
        
        # Create test supplier
        self.supplier = Supplier.objects.create(
            name='Test Supplier',
            supplier_type='local',
            currency='USD',
            phone_number='1234567890',
            email='supplier@example.com',
            delivery_terms='FOB',
            remarks='Test supplier remarks'
        )
        
        # Create test brand
        self.brand = Brand.objects.create(
            name='Test Brand',
            made_in='USA',
            show_made_in=True,
            remarks='Test brand remarks'
        )
        
        # Create test categories
        self.category = Category.objects.create(
            name='Electronics',
            parent=None
        )
        
        self.subcategory = Category.objects.create(
            name='Laptops',
            parent=self.category
        )
        
        self.sub_level_category = Category.objects.create(
            name='Gaming Laptops',
            parent=self.subcategory
        )
        
        # Create test inventory
        self.inventory = Inventory.objects.create(
            item_code='TEST001',
            cip_code='CIP001',
            product_name='Test Product',
            status='active',
            supplier=self.supplier,
            brand=self.brand,
            product_tagging='never_sold',
            audit_status=False,
            category=self.category,
            subcategory=self.subcategory,
            sub_level_category=self.sub_level_category,
            created_by=self.user,
            last_modified_by=self.user
        )
        
        # Set up API client
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)
        
        # Define common test data
        self.general_data = {
            'item_code': 'TEST002',
            'cip_code': 'CIP002',
            'product_name': 'New Test Product',
            'status': 'active',
            'supplier': self.supplier.id,
            'brand': self.brand.id,
            'product_tagging': 'never_sold',
            'audit_status': False,
            'category': self.category.id,
            'subcategory': self.subcategory.id,
            'sub_level_category': self.sub_level_category.id
        }
        
        self.description_data = {
            'unit': 'pcs',
            'landed_cost_price': '100.00',
            'landed_cost_unit': 'USD',
            'packaging_amount': 10,
            'packaging_units': 'pcs',
            'packaging_package': 'Box',
            'external_description': 'Test description',
            'length': '10.5',
            'length_unit': 'cm',
            'color': 'Black',
            'width': '5.2',
            'width_unit': 'cm',
            'height': '3.1',
            'height_unit': 'cm',
            'volume': '170.5',
            'volume_unit': 'cm³',
            'materials': 'Plastic, Metal',
            'list_price_currency': 'USD',
            'list_price': '150.00',
            'wholesale_price': '120.00',
            'remarks': 'Test remarks'
        }
        
        # Define URLs using the 'admin_api' namespace within setUp or where needed
        self.list_url = reverse('admin_api:inventory-list')
        self.detail_url = lambda pk: reverse('admin_api:inventory-detail', args=[pk])
        self.general_create_url = reverse('admin_api:inventory-general-create')
        self.general_update_url = lambda pk: reverse('admin_api:inventory-general-update', args=[pk])
        self.description_update_url = lambda pk: reverse('admin_api:inventory-description-update', args=[pk])
        self.download_template_url = reverse('admin_api:inventory-download-template')
        self.upload_url = reverse('admin_api:inventory-upload')
        # Note: The supplier list URL likely only exists in inventory_api, not admin_api
    
    def test_get_inventory_list(self):
        """Test retrieving a list of inventory items."""
        response = self.client.get(self.list_url)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(len(response.data['data']), 1)
        self.assertEqual(response.data['data'][0]['item_code'], 'TEST001')
    
    def test_get_inventory_detail(self):
        """Test retrieving a single inventory item."""
        url = self.detail_url(self.inventory.id)
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['data']['id'], self.inventory.id)
        self.assertEqual(response.data['data']['item_code'], 'TEST001')
        self.assertEqual(response.data['data']['cip_code'], 'CIP001')
        self.assertEqual(response.data['data']['product_name'], 'Test Product')
        self.assertEqual(response.data['data']['supplier_name'], 'Test Supplier')
        self.assertEqual(response.data['data']['brand_name'], 'Test Brand')
        self.assertEqual(response.data['data']['made_in'], 'USA')
    
    def test_create_inventory_general(self):
        """Test creating a new inventory item with general information."""
        url = self.general_create_url
        response = self.client.post(url, self.general_data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['data']['item_code'], 'TEST002')
        self.assertEqual(response.data['data']['cip_code'], 'CIP002')
        self.assertEqual(response.data['data']['product_name'], 'New Test Product')
        
        # Verify the inventory was created in the database
        self.assertTrue(Inventory.objects.filter(item_code='TEST002').exists())
    
    def test_update_inventory_general(self):
        """Test updating general information of an existing inventory item."""
        url = self.general_update_url(self.inventory.id)
        update_data = {
            'product_name': 'Updated Product Name',
            'status': 'inactive'
        }
        
        response = self.client.put(url, update_data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['data']['product_name'], 'Updated Product Name')
        self.assertEqual(response.data['data']['status'], 'inactive')
        
        # Verify the inventory was updated in the database
        self.inventory.refresh_from_db()
        self.assertEqual(self.inventory.product_name, 'Updated Product Name')
        self.assertEqual(self.inventory.status, 'inactive')
    
    def test_update_inventory_description(self):
        """Test updating description information of an existing inventory item."""
        url = self.description_update_url(self.inventory.id)
        
        # Create test image file
        image = SimpleUploadedFile(
            name='test_image.jpg',
            content=b'',  # Empty content for simplicity
            content_type='image/jpeg'
        )
        
        data = {
            'unit': 'pcs',
            'landed_cost_price': '100.00',
            'landed_cost_unit': 'USD',
            'packaging_amount': 1,
            'packaging_units': 'box',
            'packaging_package': 'carton',
            'external_description': 'Test description',
            'length': '10.00',
            'length_unit': 'cm',
            'color': 'Black',
            'width': '5.00',
            'width_unit': 'cm',
            'height': '2.00',
            'height_unit': 'cm',
            'volume': '100.00',
            'volume_unit': 'cm³',
            'materials': 'Plastic',
            'list_price_currency': 'USD',
            'list_price': '150.00',
            'wholesale_price': '120.00',
            'remarks': 'Test remarks'
        }
        
        # Use multipart format for file upload
        response = self.client.put(url, data, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Refresh inventory from database
        self.inventory.refresh_from_db()
        
        # Check that description fields were updated
        self.assertEqual(self.inventory.unit, 'pcs')
        self.assertEqual(self.inventory.landed_cost_price, Decimal('100.00'))
        self.assertEqual(self.inventory.color, 'Black')
        self.assertTrue(self.inventory.has_description)
    
    def test_delete_inventory(self):
        """Test deleting an inventory item."""
        url = self.detail_url(self.inventory.id)
        response = self.client.delete(url)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Verify the inventory was deleted from the database
        self.assertFalse(Inventory.objects.filter(id=self.inventory.id).exists())
    
    def test_search_inventory(self):
        """Test searching for inventory items."""
        # Create additional inventory items for testing search
        Inventory.objects.create(
            item_code='SEARCH001',
            cip_code='CIPSEARCH001',
            product_name='Searchable Product',
            status='active',
            supplier=self.supplier,
            brand=self.brand,
            product_tagging='never_sold',
            audit_status=False,
            category=self.category,
            created_by=self.user,
            last_modified_by=self.user
        )
        
        response = self.client.get(self.list_url, {'search': 'search'})
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(len(response.data['data']), 1)
        self.assertEqual(response.data['data'][0]['item_code'], 'SEARCH001')
    
    def test_filter_inventory_by_status(self):
        """Test filtering inventory items by status."""
        # Create an inactive inventory item
        Inventory.objects.create(
            item_code='INACTIVE001',
            cip_code='CIPINACTIVE001',
            product_name='Inactive Product',
            status='inactive',
            supplier=self.supplier,
            brand=self.brand,
            product_tagging='never_sold',
            audit_status=False,
            category=self.category,
            created_by=self.user,
            last_modified_by=self.user
        )
        
        response = self.client.get(self.list_url, {'status': 'inactive'})
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(len(response.data['data']), 1)
        self.assertEqual(response.data['data'][0]['item_code'], 'INACTIVE001')
    
    def test_filter_inventory_by_supplier(self):
        """Test filtering inventory items by supplier."""
        # Create a second supplier
        new_supplier = Supplier.objects.create(
            name='Another Supplier',
            supplier_type='foreign',
            currency='EURO',
            phone_number='9876543210',
            email='another@example.com',
            delivery_terms='CIF',
            remarks='Another test supplier'
        )
        
        # Create a second inventory item with the new supplier
        Inventory.objects.create(
            item_code='TEST004',
            cip_code='CIP004',
            product_name='Imported Product',
            status='active',
            supplier=new_supplier,
            brand=self.brand,
            product_tagging='never_sold',
            audit_status=False,
            category=self.category,
            subcategory=self.subcategory,
            sub_level_category=self.sub_level_category,
            created_by=self.user,
            last_modified_by=self.user
        )
        
        # Filter by the original supplier
        response = self.client.get(self.list_url + f'?supplier_id={self.supplier.id}')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(len(response.data['data']), 1)
        self.assertEqual(response.data['data'][0]['item_code'], 'TEST001')
    
    def test_filter_inventory_by_brand(self):
        """Test filtering inventory items by brand."""
        # Create a second brand
        new_brand = Brand.objects.create(
            name='Another Brand',
            made_in='China',
            show_made_in=True,
            remarks='Another test brand'
        )
        
        # Create a second inventory item with the new brand
        Inventory.objects.create(
            item_code='TEST002',
            cip_code='CIP002',
            product_name='Another Product',
            status='active',
            supplier=self.supplier,
            brand=new_brand,
            product_tagging='never_sold',
            audit_status=False,
            category=self.category,
            subcategory=self.subcategory,
            sub_level_category=self.sub_level_category,
            created_by=self.user,
            last_modified_by=self.user
        )
        
        # Filter by the original brand
        response = self.client.get(self.list_url + f'?brand_id={self.brand.id}')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(len(response.data['data']), 1)
        self.assertEqual(response.data['data'][0]['item_code'], 'TEST001')
    
    def test_filter_inventory_by_category(self):
        """Test filtering inventory items by category."""
        # Create a second category
        new_category = Category.objects.create(
            name='Furniture',
            parent=None
        )
        
        new_subcategory = Category.objects.create(
            name='Chairs',
            parent=new_category
        )
        
        # Create a second inventory item with the new category
        Inventory.objects.create(
            item_code='TEST003',
            cip_code='CIP003',
            product_name='Office Chair',
            status='active',
            supplier=self.supplier,
            brand=self.brand,
            product_tagging='never_sold',
            audit_status=False,
            category=new_category,
            subcategory=new_subcategory,
            created_by=self.user,
            last_modified_by=self.user
        )
        
        # Filter by the original category
        response = self.client.get(self.list_url + f'?category_id={self.category.id}')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(len(response.data['data']), 1)
        self.assertEqual(response.data['data'][0]['item_code'], 'TEST001')
    
    def test_validation_category_hierarchy(self):
        """Test validation of category hierarchy."""
        # Create a different category hierarchy
        different_category = Category.objects.create(
            name='Appliances',
            parent=None
        )
        
        different_subcategory = Category.objects.create(
            name='Refrigerators',
            parent=different_category
        )
        
        # Try to create an inventory with mismatched category hierarchy
        url = self.general_create_url
        data = {
            'item_code': 'TEST005',
            'cip_code': 'CIP005',
            'product_name': 'Invalid Product',
            'status': 'active',
            'supplier': self.supplier.id,
            'brand': self.brand.id,
            'product_tagging': 'never_sold',
            'audit_status': False,
            'category': self.category.id,
            'subcategory': different_subcategory.id  # This should fail validation
        }
        
        response = self.client.post(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('subcategory', response.data['errors'])
    
    def test_download_template(self):
        """Test downloading the inventory template."""
        response = self.client.get(self.download_template_url)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertEqual(
            response['Content-Disposition'],
            'attachment; filename=inventory_template.xlsx'
        )
        
        # Verify the template content
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check headers
        self.assertEqual(ws.cell(row=1, column=1).value, 'Item Code*')
        self.assertEqual(ws.cell(row=1, column=2).value, 'CIP Code*')
        self.assertEqual(ws.cell(row=1, column=3).value, 'Product Name*')
        self.assertEqual(ws.cell(row=1, column=4).value, 'Status*')
    
    def test_upload_inventory(self):
        """Test uploading inventory data."""
        url = self.upload_url
        
        # Create a test Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add headers - update these to match what your view expects
        headers = [
            'Item Code*', 'CIP Code*', 'Product Name*', 'Status*', 'Supplier ID*', 'Brand ID*',
            'Product Tagging*', 'Audit Status*', 'Category ID*', 'Subcategory ID', 'Sub Level Category ID'
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).value = header
        
        # Add a row of data - make sure the values match the expected format
        data = [
            'UPLOAD001', 'CIPUPLOAD001', 'Uploaded Product', 'active', 
            str(self.supplier.id), str(self.brand.id), 'never_sold', 'False',
            str(self.category.id), str(self.subcategory.id), str(self.sub_level_category.id)
        ]
        for col_num, value in enumerate(data, 1):
            ws.cell(row=2, column=col_num).value = value
        
        # Save to a BytesIO object
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create a SimpleUploadedFile from the BytesIO object
        upload_file = SimpleUploadedFile(
            'test_upload.xlsx',
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        response = self.client.post(url, {'file': upload_file}, format='multipart')
        
        # If your view is returning errors, print them for debugging
        if response.data['data']['success_count'] == 0 and 'errors' in response.data['data']:
            print("Upload errors:", response.data['data']['errors'])
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['data']['total_rows'], 1)
        self.assertEqual(response.data['data']['success_count'], 1)
        self.assertEqual(response.data['data']['error_count'], 0)
        
        # Verify the inventory was created in the database
        self.assertTrue(Inventory.objects.filter(item_code='UPLOAD001').exists())
    
    def test_upload_inventory_validation_errors(self):
        """Test validation errors when uploading inventory data."""
        url = self.upload_url
        
        # Create a test Excel file with invalid data
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add headers
        headers = [
            'Item Code*', 'CIP Code*', 'Product Name*', 'Status*', 'Supplier ID*', 'Brand ID*',
            'Product Tagging*', 'Audit Status*', 'Category ID*', 'Subcategory ID', 'Sub Level Category ID'
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).value = header
        
        # Add a row with invalid data (invalid status, non-existent supplier)
        data = [
            'INVALID001', 'CIPINVALID001', 'Invalid Product', 'pending',  # Invalid status
            '9999',  # Non-existent supplier
            str(self.brand.id), 'never_sold', 'False',
            str(self.category.id), '', ''
        ]
        for col_num, value in enumerate(data, 1):
            ws.cell(row=2, column=col_num).value = value
        
        # Save to a BytesIO object
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create a SimpleUploadedFile from the BytesIO object
        upload_file = SimpleUploadedFile(
            'invalid_upload.xlsx',
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        response = self.client.post(url, {'file': upload_file}, format='multipart')
        
        # Your view returns 200 even for validation errors
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Check that the response contains error information
        self.assertEqual(response.data['data']['success_count'], 0)
        self.assertEqual(response.data['data']['error_count'], 1)
        self.assertTrue(len(response.data['data']['errors']) > 0)
        
        # Check specific error messages
        errors = response.data['data']['errors'][0]['errors']
        self.assertIn('status', errors)
        self.assertIn('supplier', errors)
        
        # Verify the inventory was not created in the database
        self.assertFalse(Inventory.objects.filter(item_code='INVALID001').exists())
    
    def test_upload_inventory_duplicate_item_code(self):
        """Test uploading inventory with duplicate item code."""
        url = self.upload_url
        
        # Create a test Excel file with duplicate item code
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add headers
        headers = [
            'Item Code*', 'CIP Code*', 'Product Name*', 'Status*', 'Supplier ID*', 'Brand ID*',
            'Product Tagging*', 'Audit Status*', 'Category ID*', 'Subcategory ID', 'Sub Level Category ID'
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).value = header
        
        # Add a row with duplicate item code
        data = [
            'TEST001',  # Already exists from setUp
            'CIPDUP001',  # New CIP code
            'Duplicate Product', 'active', 
            str(self.supplier.id), str(self.brand.id), 'never_sold', 'False',
            str(self.category.id), str(self.subcategory.id), str(self.sub_level_category.id)
        ]
        for col_num, value in enumerate(data, 1):
            ws.cell(row=2, column=col_num).value = value
        
        # Save to a BytesIO object
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create a SimpleUploadedFile from the BytesIO object
        upload_file = SimpleUploadedFile(
            'duplicate_upload.xlsx',
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        response = self.client.post(url, {'file': upload_file}, format='multipart')
        
        # Your view returns 200 even for validation errors
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Check that the response contains error information
        self.assertEqual(response.data['data']['success_count'], 0)
        self.assertEqual(response.data['data']['error_count'], 1)
        self.assertTrue(len(response.data['data']['errors']) > 0)
        
        # Check specific error messages
        errors = response.data['data']['errors'][0]['errors']
        self.assertIn('item_code', errors)