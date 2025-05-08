import io
from decimal import Decimal
from PIL import Image  # Make sure to import PIL (Pillow)

import openpyxl
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.conf import settings

from rest_framework import status
from rest_framework.test import APITestCase
from rest_framework_simplejwt.tokens import RefreshToken

# Reuse models from admin_api
from admin_api.models import Inventory, Supplier, Brand, Category

User = get_user_model()


class InventoryAPITests(APITestCase):
    """Tests for the Inventory API endpoints in inventory_api."""

    @classmethod
    def setUpTestData(cls):
        """Set up data for the whole test class."""
        cls.user = User.objects.create_user(
            username='testuser',
            password='password123',
            email='test@example.com',
            role='admin',  # Assuming role is needed
            is_staff=True,
            is_active=True
        )
        cls.supplier = Supplier.objects.create(
            name='Test Supplier',
            supplier_type='local',
            currency='PHP',
            phone_number='1234567890',
            email='supplier@example.com'
        )
        cls.brand = Brand.objects.create(name='Test Brand')
        cls.category = Category.objects.create(name='Test Category')
        cls.subcategory = Category.objects.create(name='Test SubCategory', parent=cls.category)
        cls.sub_level_category = Category.objects.create(name='Test SubLevel', parent=cls.subcategory)

        # Create multiple inventory items for pagination/filtering tests
        cls.inventory1 = Inventory.objects.create(
            item_code='TEST001',
            cip_code='CIP001',
            product_name='Test Product One',
            status='active',
            supplier=cls.supplier,
            brand=cls.brand,
            category=cls.category,
            subcategory=cls.subcategory,
            sub_level_category=cls.sub_level_category,
            created_by=cls.user,
            last_modified_by=cls.user,
            has_description=False
        )
        cls.inventory2 = Inventory.objects.create(
            item_code='TEST002',
            cip_code='CIP002',
            product_name='Test Product Two',
            status='inactive',
            supplier=cls.supplier,
            brand=cls.brand,
            category=cls.category,
            created_by=cls.user,
            last_modified_by=cls.user,
            has_description=True,
            unit='pcs',
            color='Red'
        )
        # Add more items if needed for thorough pagination testing

        # URLs - Use the 'inventory_api' namespace
        cls.list_url = reverse('inventory_api:inventory-list')
        cls.detail_url = lambda pk: reverse('inventory_api:inventory-detail', args=[pk])
        cls.general_create_url = reverse('inventory_api:inventory-general-create')
        cls.general_update_url = lambda pk: reverse('inventory_api:inventory-general-update', args=[pk])
        cls.description_update_url = lambda pk: reverse('inventory_api:inventory-description-update', args=[pk])
        cls.download_template_url = reverse('inventory_api:inventory-download-template')
        cls.upload_url = reverse('inventory_api:inventory-upload')
        cls.supplier_list_url = reverse('inventory_api:supplier-list') # Namespace added

    def setUp(self):
        """Set up for each test method."""
        # Authenticate the client
        refresh = RefreshToken.for_user(self.user)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {refresh.access_token}')

        # Data for POST/PUT requests
        self.general_data = {
            'item_code': 'NEW001',
            'cip_code': 'CIPNEW001',
            'product_name': 'New Test Product',
            'status': 'active',
            'supplier': self.supplier.id,
            'brand': self.brand.id,
            'product_tagging': 'never_sold',
            'audit_status': False,
            'category': self.category.id,
            'subcategory': self.subcategory.id,
            'sub_level_category': self.sub_level_category.id,
        }
        self.description_data = {
            'unit': 'kg',
            'landed_cost_price': '50.00',
            'landed_cost_unit': 'PHP',
            'packaging_amount': 5,
            'packaging_units': 'bags',
            'packaging_package': 'Sack',
            'external_description': 'Updated description',
            'length': '20.0',
            'length_unit': 'm',
            'color': 'Blue',
            'width': '15.0',
            'width_unit': 'm',
            'height': '10.0',
            'height_unit': 'm',
            'volume': '3000.0',
            'volume_unit': 'm³',
            'materials': 'Fabric',
            'pattern': 'Pattern 1',
            'list_price_currency': 'PHP',
            'list_price': '75.00',
            'wholesale_price': '60.00',
            'remarks': 'Updated remarks'
        }

    # --- Authentication Tests ---
    def test_list_inventory_unauthenticated(self):
        """Test accessing inventory list without authentication fails."""
        self.client.credentials() # Clear authentication
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    # --- List View Tests (GET /api/inventory/) ---
    def test_get_inventory_list_success(self):
        """Test retrieving a list of inventory items successfully."""
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['meta']['pagination']['count'], 2)
        self.assertIsInstance(response.data['data'], list)
        # Check default pagination size
        self.assertEqual(len(response.data['data']), min(settings.REST_FRAMEWORK['PAGE_SIZE'], 2))

    def test_get_inventory_list_pagination_custom_size(self):
        """Test retrieving inventory list with custom page size."""
        page_size = 1 # Request an invalid page size according to view logic
        response = self.client.get(self.list_url, {'page_size': page_size})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        total_items = response.data['meta']['pagination']['count']
        default_page_size = settings.REST_FRAMEWORK.get('PAGE_SIZE', 10) # Get default size

        # Assert that the number of items returned is the default size,
        # capped by the total number of items available.
        expected_size = min(default_page_size, total_items)
        self.assertEqual(len(response.data['data']), expected_size)

        # Check next/previous links based on the effective page size (default)
        if total_items > default_page_size:
             self.assertIsNotNone(response.data['meta']['pagination']['next'])
        else:
             self.assertIsNone(response.data['meta']['pagination']['next'])

    def test_get_inventory_list_pagination_invalid_size(self):
        """Test retrieving inventory list with invalid page size uses default."""
        response = self.client.get(self.list_url, {'page_size': 999}) # Assume 999 > max_page_size
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        # Should default to PAGE_SIZE or max_page_size if 999 is allowed but capped
        # Check against the actual max_page_size set in the view or default PAGE_SIZE
        expected_size = min(settings.REST_FRAMEWORK['PAGE_SIZE'], 100) # Assuming max_page_size=100
        self.assertEqual(len(response.data['data']), min(expected_size, 2)) # Capped by total items

    def test_get_inventory_list_pagination_non_allowed_size(self):
        """Test retrieving inventory list with a non-allowed page size uses default."""
        response = self.client.get(self.list_url, {'page_size': 15}) # 15 is not in [10, 25, 50, 100]
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(len(response.data['data']), min(settings.REST_FRAMEWORK['PAGE_SIZE'], 2))

    def test_search_inventory_item_code(self):
        """Test searching inventory by item code."""
        response = self.client.get(self.list_url, {'search': 'TEST001'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['meta']['pagination']['count'], 1)
        self.assertEqual(response.data['data'][0]['item_code'], 'TEST001')

    def test_search_inventory_product_name(self):
        """Test searching inventory by product name."""
        response = self.client.get(self.list_url, {'search': 'Product Two'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['meta']['pagination']['count'], 1)
        self.assertEqual(response.data['data'][0]['product_name'], 'Test Product Two')

    def test_filter_inventory_by_status(self):
        """Test filtering inventory by status."""
        response = self.client.get(self.list_url, {'status': 'inactive'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['meta']['pagination']['count'], 1)
        self.assertEqual(response.data['data'][0]['status'], 'inactive')
        self.assertEqual(response.data['data'][0]['item_code'], 'TEST002')

    def test_filter_inventory_by_supplier(self):
        """Test filtering inventory by supplier ID."""
        response = self.client.get(self.list_url, {'supplier_id': self.supplier.id})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['meta']['pagination']['count'], 2) # Both items have this supplier

    def test_filter_inventory_by_brand(self):
        """Test filtering inventory by brand ID."""
        response = self.client.get(self.list_url, {'brand_id': self.brand.id})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['meta']['pagination']['count'], 2) # Both items have this brand

    def test_filter_inventory_by_category(self):
        """Test filtering inventory by category ID."""
        response = self.client.get(self.list_url, {'category_id': self.category.id})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['meta']['pagination']['count'], 2) # Both items have this category

    def test_filter_inventory_by_specific_item_code(self):
        """Test filtering inventory by specific item_code parameter."""
        response = self.client.get(self.list_url, {'item_code': 'TEST002'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['meta']['pagination']['count'], 1)
        self.assertEqual(response.data['data'][0]['item_code'], 'TEST002')

    def test_filter_inventory_by_specific_cip_code(self):
        """Test filtering inventory by specific cip_code parameter."""
        response = self.client.get(self.list_url, {'cip_code': 'CIP001'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['meta']['pagination']['count'], 1)
        self.assertEqual(response.data['data'][0]['cip_code'], 'CIP001')

    def test_filter_inventory_by_specific_product_name(self):
        """Test filtering inventory by specific product_name parameter."""
        response = self.client.get(self.list_url, {'product_name': 'Test Product One'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['meta']['pagination']['count'], 1)
        self.assertEqual(response.data['data'][0]['product_name'], 'Test Product One')

    def test_sort_inventory_by_item_code_desc(self):
        """Test sorting inventory by item_code descending."""
        response = self.client.get(self.list_url, {'sort_by': 'item_code', 'sort_direction': 'desc'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['data'][0]['item_code'], 'TEST002')
        self.assertEqual(response.data['data'][1]['item_code'], 'TEST001')

    def test_sort_inventory_by_product_name_asc(self):
        """Test sorting inventory by product_name ascending."""
        response = self.client.get(self.list_url, {'sort_by': 'product_name', 'sort_direction': 'asc'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['data'][0]['product_name'], 'Test Product One')
        self.assertEqual(response.data['data'][1]['product_name'], 'Test Product Two')

    # --- Detail View Tests (GET /api/inventory/{pk}/) ---
    def test_get_inventory_detail_success(self):
        """Test retrieving a single inventory item successfully."""
        url = self.detail_url(self.inventory1.id)
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['data']['id'], self.inventory1.id)
        self.assertEqual(response.data['data']['item_code'], self.inventory1.item_code)
        self.assertEqual(response.data['data']['supplier'], self.supplier.id) # Check ID
        self.assertEqual(response.data['data']['supplier_name'], self.supplier.name) # Check name

    def test_get_inventory_detail_not_found(self):
        """Test retrieving a non-existent inventory item returns 404."""
        url = self.detail_url(9999) # Non-existent ID
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
        self.assertIn('detail', response.data)

    # --- General Create Tests (POST /api/inventory/general/) ---
    def test_create_inventory_general_success(self):
        """Test creating a new inventory item with general info successfully."""
        initial_count = Inventory.objects.count()
        response = self.client.post(self.general_create_url, self.general_data, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertTrue(response.data['success'])
        self.assertEqual(Inventory.objects.count(), initial_count + 1)
        self.assertEqual(response.data['data']['item_code'], self.general_data['item_code'])
        self.assertEqual(response.data['data']['supplier'], self.supplier.id)
        self.assertFalse(response.data['data']['has_description']) # Should be false initially

    def test_create_inventory_general_missing_data(self):
        """Test creating inventory with missing required general info fails."""
        data = self.general_data.copy()
        del data['item_code'] # Remove required field
        response = self.client.post(self.general_create_url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('item_code', response.data['errors'])

    def test_create_inventory_general_invalid_supplier(self):
        """Test creating inventory with invalid supplier ID fails."""
        data = self.general_data.copy()
        data['supplier'] = 9999 # Invalid ID
        response = self.client.post(self.general_create_url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('supplier', response.data['errors'])

    # --- General Update Tests (PUT /api/inventory/{pk}/general/) ---
    def test_update_inventory_general_success(self):
        """Test updating general info of an existing inventory item successfully."""
        url = self.general_update_url(self.inventory1.id)
        data = {
            'product_name': 'Updated Product Name',
            'status': 'inactive',
            'item_code': 'TEST001-UPDATED' # Test updating unique fields too
        }
        response = self.client.put(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.inventory1.refresh_from_db()
        self.assertEqual(self.inventory1.product_name, data['product_name'])
        self.assertEqual(self.inventory1.status, data['status'])
        self.assertEqual(self.inventory1.item_code, data['item_code'])

    def test_update_inventory_general_partial_success(self):
        """Test partially updating general info."""
        url = self.general_update_url(self.inventory1.id)
        data = {'status': 'inactive'} # Only update status
        response = self.client.put(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.inventory1.refresh_from_db()
        self.assertEqual(self.inventory1.status, data['status'])
        self.assertEqual(self.inventory1.product_name, 'Test Product One') # Name should be unchanged

    def test_update_inventory_general_invalid_data(self):
        """Test updating general info with invalid data fails."""
        url = self.general_update_url(self.inventory1.id)
        data = {'supplier': 9999} # Invalid supplier ID
        response = self.client.put(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('supplier', response.data['errors'])

    def test_update_inventory_general_not_found(self):
        """Test updating general info for a non-existent item fails."""
        url = self.general_update_url(9999)
        data = {'status': 'inactive'}
        response = self.client.put(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
        self.assertIn('detail', response.data) 

    # --- Description Update Tests (PUT /api/inventory/{pk}/description/) ---
    def test_update_inventory_description_success(self):
        """Test updating description info successfully."""
        url = self.description_update_url(self.inventory1.id)
        response = self.client.put(url, self.description_data, format='json') # Use JSON for non-file data
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.inventory1.refresh_from_db()
        self.assertEqual(self.inventory1.unit, self.description_data['unit'])
        self.assertEqual(self.inventory1.color, self.description_data['color'])
        self.assertEqual(self.inventory1.landed_cost_price, Decimal(self.description_data['landed_cost_price']))
        self.assertTrue(self.inventory1.has_description) # Flag should be set

    def test_update_inventory_description_with_photo(self):
        """Test updating description info including a photo upload."""
        url = self.description_update_url(self.inventory1.id)
        
        # Create a proper test image using PIL
        image_io = io.BytesIO()
        image = Image.new('RGB', (100, 100), color='red')
        image.save(image_io, format='JPEG')
        image_io.seek(0)
        
        image_file = SimpleUploadedFile(
            name='test_image.jpg',
            content=image_io.getvalue(),
            content_type='image/jpeg'
        )
        
        data = self.description_data.copy()
        data['photo'] = image_file

        response = self.client.put(url, data, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.inventory1.refresh_from_db()
        self.assertTrue(self.inventory1.has_description)
        self.assertTrue(self.inventory1.photo.name.endswith('.jpg'))  # Just check extension
        # Clean up the uploaded file after test if necessary

    def test_update_inventory_description_invalid_data(self):
        """Test updating description info with invalid data fails."""
        url = self.description_update_url(self.inventory1.id)
        data = self.description_data.copy()
        data['landed_cost_price'] = 'not-a-number'
        response = self.client.put(url, data, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('landed_cost_price', response.data['errors'])

    def test_update_inventory_description_not_found(self):
        """Test updating description info for a non-existent item fails."""
        url = self.description_update_url(9999)
        response = self.client.put(url, self.description_data, format='json')
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
        self.assertIn('detail', response.data)

    # --- Template Download Test (GET /api/inventory/download-template/) ---
    def test_download_template_success(self):
        """Test downloading the inventory template successfully."""
        response = self.client.get(self.download_template_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertTrue(
            response['Content-Disposition'].startswith('attachment; filename=')
        )
        # Optional: Deeper inspection of the file content
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, 'Item Code*') # Check first header

    # --- Upload Test (POST /api/inventory/upload/) ---
    def test_upload_inventory_success(self):
        """Test uploading inventory data successfully."""
        initial_count = Inventory.objects.count()
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [ # Match headers expected by InventoryUploadView
            'Item Code*', 'CIP Code*', 'Product Name*', 'Status*', 'Supplier ID*', 'Brand ID*',
            'Product Tagging*', 'Audit Status*', 'Category ID*', 'Subcategory ID', 'Sub Level Category ID'
        ]
        ws.append(headers)
        data_row = [
            'UPLOAD001', 'CIPUPLOAD001', 'Uploaded Product', 'active',
            str(self.supplier.id), str(self.brand.id), 'never_sold', 'False',
            str(self.category.id), str(self.subcategory.id), str(self.sub_level_category.id)
        ]
        ws.append(data_row)

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        upload_file = SimpleUploadedFile(
            'test_upload.xlsx',
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response = self.client.post(self.upload_url, {'file': upload_file}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['data']['total_rows'], 1)
        self.assertEqual(response.data['data']['success_count'], 1)
        self.assertEqual(response.data['data']['error_count'], 0)
        self.assertEqual(Inventory.objects.count(), initial_count + 1)
        self.assertTrue(Inventory.objects.filter(item_code='UPLOAD001').exists())

    def test_upload_inventory_validation_errors(self):
        """Test uploading inventory data with validation errors."""
        initial_count = Inventory.objects.count()
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [ # Match headers
            'Item Code*', 'CIP Code*', 'Product Name*', 'Status*', 'Supplier ID*', 'Brand ID*',
            'Product Tagging*', 'Audit Status*', 'Category ID*', 'Subcategory ID', 'Sub Level Category ID'
        ]
        ws.append(headers)
        data_row = [ # Missing required Product Name, invalid Supplier ID
            'UPLOAD002', 'CIPUPLOAD002', '', 'active',
            '9999', str(self.brand.id), 'never_sold', 'False',
            str(self.category.id), '', ''
        ]
        ws.append(data_row)

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        upload_file = SimpleUploadedFile(
            'test_upload_errors.xlsx',
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response = self.client.post(self.upload_url, {'file': upload_file}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK) # Upload itself is OK
        self.assertTrue(response.data['success']) # API call success
        self.assertEqual(response.data['data']['total_rows'], 1)
        self.assertEqual(response.data['data']['success_count'], 0)
        self.assertEqual(response.data['data']['error_count'], 1)
        self.assertIn('errors', response.data['data'])
        self.assertEqual(len(response.data['data']['errors']), 1)
        self.assertEqual(response.data['data']['errors'][0]['row'], 2) # Row number in Excel
        self.assertIn('product_name', response.data['data']['errors'][0]['errors'])
        self.assertIn('supplier', response.data['data']['errors'][0]['errors'])
        self.assertEqual(Inventory.objects.count(), initial_count) # No inventory created

    def test_upload_inventory_invalid_file(self):
        """Test uploading an invalid file type."""
        upload_file = SimpleUploadedFile(
            'test.txt',
            b'this is not excel',
            content_type='text/plain'
        )
        response = self.client.post(self.upload_url, {'file': upload_file}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('file', response.data['errors']) # Should indicate file format error

    # --- Supplier List Test (GET /api/inventory/suppliers/) ---
    def test_get_supplier_list_success(self):
        """Test retrieving the list of suppliers successfully."""
        # Create another supplier for testing list length
        another_supplier = Supplier.objects.create(
            name='Another Supplier', supplier_type='foreign', currency='USD',
            phone_number='0987654321', email='another@example.com'
        )
        response = self.client.get(self.supplier_list_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertIsInstance(response.data['data'], list)
        self.assertEqual(len(response.data['data']), 2) # Test Supplier + Another Supplier

        # *** FIX: Check that both expected IDs are present using sets ***
        response_ids = {item['id'] for item in response.data['data']}
        expected_ids = {self.supplier.id, another_supplier.id}
        self.assertEqual(response_ids, expected_ids)

        # Optionally, check structure of one item (find it first)
        first_supplier_data = next((item for item in response.data['data'] if item['id'] == self.supplier.id), None)
        self.assertIsNotNone(first_supplier_data)
        self.assertIn('id', first_supplier_data)
        self.assertIn('name', first_supplier_data)
        self.assertEqual(first_supplier_data['name'], self.supplier.name)
        # Ensure only id and name are present as per the view's serializer fields
        self.assertEqual(len(first_supplier_data.keys()), 2)

    def test_get_supplier_list_unauthenticated(self):
        """Test accessing supplier list without authentication fails."""
        self.client.credentials() # Clear authentication
        response = self.client.get(self.supplier_list_url)
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
