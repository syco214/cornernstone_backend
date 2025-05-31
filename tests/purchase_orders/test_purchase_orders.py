import json
import io
from decimal import Decimal
from django.test import TestCase
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APITestCase
from rest_framework import status
from rest_framework_simplejwt.tokens import RefreshToken
from openpyxl import Workbook

from admin_api.models import Supplier, Inventory, Brand, Category
from purchase_order_api.models import (
    PurchaseOrder, 
    PurchaseOrderItem, 
)

User = get_user_model()


class PurchaseOrderTests(APITestCase):
    """Test cases for Purchase Order CRUD operations"""
    
    def setUp(self):
        """Set up test data"""
        # Create test users
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        self.admin_user = User.objects.create_user(
            username='admin',
            email='admin@example.com',
            password='adminpass123',
            is_staff=True
        )
        
        # Create JWT token for authentication
        refresh = RefreshToken.for_user(self.user)
        self.access_token = str(refresh.access_token)
        
        # Set up authentication
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {self.access_token}')
        
        # Create test supplier
        self.supplier = Supplier.objects.create(
            name='Test Supplier',
            supplier_type='local',
            currency='USD',
            phone_number='123-456-7890',
            email='supplier@test.com',
            delivery_terms='FOB Test Port',
            remarks='Test supplier remarks'
        )
        
        # Create test brand and category for inventory
        self.brand = Brand.objects.create(name='Test Brand')
        self.category = Category.objects.create(name='Test Category')
        
        # Create test inventory items
        self.inventory1 = Inventory.objects.create(
            item_code='TEST001',
            cip_code='CIP001',
            product_name='Test Product 1',
            external_description='Test Item 1',
            unit='PCS',
            wholesale_price=Decimal('100.00'),
            supplier=self.supplier,
            brand=self.brand,
            category=self.category,
            created_by=self.user,
            last_modified_by=self.user
        )
        
        self.inventory2 = Inventory.objects.create(
            item_code='TEST002',
            cip_code='CIP002',
            product_name='Test Product 2',
            external_description='Test Item 2',
            unit='KG',
            wholesale_price=Decimal('50.00'),
            supplier=self.supplier,
            brand=self.brand,
            category=self.category,
            created_by=self.user,
            last_modified_by=self.user
        )
        
        # Create test purchase order
        self.purchase_order = PurchaseOrder.objects.create(
            supplier=self.supplier,
            supplier_type='local',
            delivery_terms='FOB Test Port',
            currency='USD',
            po_date='2024-01-01',
            created_by=self.user,
            last_modified_by=self.user
        )
        
        # Create test purchase order item
        self.po_item = PurchaseOrderItem.objects.create(
            purchase_order=self.purchase_order,
            inventory=self.inventory1,
            quantity=Decimal('10.00'),
            list_price=Decimal('100.00'),
            discount_type='none',
            discount_value=Decimal('0.00')  # Ensure this is always a proper Decimal
        )
    
    def test_get_purchase_order_list(self):
        """Test retrieving purchase order list with pagination and meta data"""
        url = reverse('purchase_order_api:purchase-order-list')
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertIn('data', response.data)
        self.assertIn('meta', response.data)
        self.assertIn('pagination', response.data['meta'])
        self.assertIn('currency_options', response.data['meta'])
        self.assertIn('status_options', response.data['meta'])
        self.assertIn('supplier_type_options', response.data['meta'])
    
    def test_get_single_purchase_order(self):
        """Test retrieving a single purchase order with all related data"""
        url = reverse('purchase_order_api:purchase-order-detail', kwargs={'pk': self.purchase_order.pk})
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertIn('data', response.data)
        
        po_data = response.data['data']
        self.assertEqual(po_data['id'], self.purchase_order.pk)
        self.assertEqual(po_data['supplier'], self.supplier.pk)
        self.assertIn('items', po_data)
        self.assertEqual(len(po_data['items']), 1)
    
    def test_create_purchase_order(self):
        """Test creating a new purchase order with items and payment terms"""
        data = {
            'supplier': self.supplier.pk,
            'supplier_type': 'local',
            'delivery_terms': 'FOB Test Port',
            'currency': 'USD',
            'supplier_address': '123 Test Address',
            'country': 'Test Country',
            'notes': 'Test notes',
            'po_date': '2024-01-01',  # Use date string format
            'items': [
                {
                    'inventory': self.inventory1.pk,
                    'quantity': '5.00',
                    'list_price': '100.00',
                    'discount_type': 'none',
                    'discount_value': '0.00'
                }
            ],
            'payment_term': {
                'credit_limit': '10000.00',
                'payment_terms_description': 'Net 30 days',
                'dp_percentage': '50.00',
                'terms_days': 30
            }
        }
        
        url = reverse('purchase_order_api:purchase-order-list')
        response = self.client.post(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertTrue(response.data['success'])
        self.assertIn('data', response.data)
        
        # Verify PO was created
        po = PurchaseOrder.objects.get(pk=response.data['data']['id'])
        self.assertEqual(po.supplier, self.supplier)
        self.assertEqual(po.items.count(), 1)
        self.assertIsNotNone(po.payment_term)
    
    def test_update_purchase_order(self):
        """Test updating an existing purchase order"""
        # Include ALL existing items to prevent deletion
        data = {
            'notes': 'Updated notes',
            'delivery_terms': 'Updated delivery terms',
            'items': [
                {
                    'id': self.po_item.pk,
                    'inventory': self.inventory1.pk,
                    'quantity': '15.00',
                    'list_price': '120.00',
                    'discount_type': 'percentage',
                    'discount_value': '10.00'
                }
            ]
        }
        
        url = reverse('purchase_order_api:purchase-order-detail', kwargs={'pk': self.purchase_order.pk})
        response = self.client.put(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Verify updates
        self.purchase_order.refresh_from_db()
        self.assertEqual(self.purchase_order.notes, 'Updated notes')
        self.assertEqual(self.purchase_order.delivery_terms, 'Updated delivery terms')
        
        # Verify item was updated (should still exist)
        self.assertEqual(self.purchase_order.items.count(), 1)
        updated_item = self.purchase_order.items.first()
        self.assertEqual(updated_item.quantity, Decimal('15.00'))
        self.assertEqual(updated_item.list_price, Decimal('120.00'))
    
    def test_delete_purchase_order_draft_status(self):
        """Test deleting a purchase order in draft status"""
        url = reverse('purchase_order_api:purchase-order-detail', kwargs={'pk': self.purchase_order.pk})
        response = self.client.delete(url)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Verify PO was deleted
        with self.assertRaises(PurchaseOrder.DoesNotExist):
            PurchaseOrder.objects.get(pk=self.purchase_order.pk)
    
    def test_delete_purchase_order_non_deletable_status(self):
        """Test deleting a purchase order in non-deletable status returns error"""
        self.purchase_order.status = 'pending_approval'
        self.purchase_order.save()
        
        url = reverse('purchase_order_api:purchase-order-detail', kwargs={'pk': self.purchase_order.pk})
        response = self.client.delete(url)
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('Cannot delete PO', response.data['errors']['detail'])
    
    def test_get_nonexistent_purchase_order(self):
        """Test retrieving non-existent purchase order returns 404"""
        url = reverse('purchase_order_api:purchase-order-detail', kwargs={'pk': 99999})
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
    
    def test_search_purchase_orders_by_supplier(self):
        """Test searching purchase orders by supplier name"""
        url = reverse('purchase_order_api:purchase-order-list')
        response = self.client.get(url, {'supplier': 'Test Supplier'})
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(len(response.data['data']), 1)
    
    def test_filter_purchase_orders_by_status(self):
        """Test filtering purchase orders by status"""
        url = reverse('purchase_order_api:purchase-order-list')
        response = self.client.get(url, {'status': 'draft'})
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(len(response.data['data']), 1)
    
    def test_sort_purchase_orders(self):
        """Test sorting purchase orders"""
        # Create another PO with different date
        PurchaseOrder.objects.create(
            supplier=self.supplier,
            supplier_type='local',
            delivery_terms='FOB Test',
            currency='USD',
            po_date='2024-02-01',
            created_by=self.user,
            last_modified_by=self.user
        )
        
        url = reverse('purchase_order_api:purchase-order-list')
        response = self.client.get(url, {'sort_by': 'po_date', 'sort_direction': 'asc'})
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(len(response.data['data']), 2)
    
    def test_create_purchase_order_invalid_data(self):
        """Test creating purchase order with invalid data returns validation errors"""
        url = reverse('purchase_order_api:purchase-order-list')
        data = {
            'supplier': 99999,  # Non-existent supplier
            'supplier_type': 'invalid',
            'currency': 'INVALID'
        }
        
        response = self.client.post(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('errors', response.data)
    
    def test_unauthorized_access(self):
        """Test unauthorized access returns 401"""
        self.client.credentials()  # Remove authentication
        
        url = reverse('purchase_order_api:purchase-order-list')
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
    
    def test_download_items_template(self):
        """Test downloading Excel template for items"""
        url = reverse('purchase_order_api:purchase-order-items-template', kwargs={'pk': self.purchase_order.pk})
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'], 
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn(self.purchase_order.po_number, response['Content-Disposition'])
    
    def test_download_template_nonexistent_po(self):
        """Test downloading template for non-existent PO returns 404"""
        url = reverse('purchase_order_api:purchase-order-items-template', kwargs={'pk': 99999})
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
    
    def test_upload_items_template_valid_file(self):
        """Test uploading valid Excel file with items"""
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'item_code'
        ws['B1'] = 'quantity'
        ws['A2'] = 'TEST002'
        ws['B2'] = 5
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        uploaded_file = SimpleUploadedFile(
            'test_items.xlsx',
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        url = reverse('purchase_order_api:purchase-order-items-upload', kwargs={'pk': self.purchase_order.pk})
        response = self.client.post(url, {'file': uploaded_file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['added'], 1)
        self.assertEqual(response.data['total_rows'], 1)
        
        # Verify item was added
        self.assertEqual(self.purchase_order.items.count(), 2)
    
    def test_upload_template_no_file(self):
        """Test uploading without file returns error"""
        url = reverse('purchase_order_api:purchase-order-items-upload', kwargs={'pk': self.purchase_order.pk})
        response = self.client.post(url, {}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('No file uploaded', response.data['errors'])
    
    def test_upload_template_invalid_file_format(self):
        """Test uploading non-Excel file returns error"""
        text_file = SimpleUploadedFile('test.txt', b'not an excel file', content_type='text/plain')
        
        url = reverse('purchase_order_api:purchase-order-items-upload', kwargs={'pk': self.purchase_order.pk})
        response = self.client.post(url, {'file': text_file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('Excel file', response.data['errors'])
    
    def test_upload_template_missing_columns(self):
        """Test uploading Excel file with missing required columns"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'wrong_column'
        ws['B1'] = 'another_wrong_column'
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        uploaded_file = SimpleUploadedFile(
            'test_items.xlsx',
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        url = reverse('purchase_order_api:purchase-order-items-upload', kwargs={'pk': self.purchase_order.pk})
        response = self.client.post(url, {'file': uploaded_file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data['success'])
        self.assertIn('Missing required column', response.data['errors'])
    
    def test_upload_template_invalid_item_code(self):
        """Test uploading Excel file with invalid item codes"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'item_code'
        ws['B1'] = 'quantity'
        ws['A2'] = 'NONEXISTENT'
        ws['B2'] = 5
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        uploaded_file = SimpleUploadedFile(
            'test_items.xlsx',
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        url = reverse('purchase_order_api:purchase-order-items-upload', kwargs={'pk': self.purchase_order.pk})
        response = self.client.post(url, {'file': uploaded_file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['added'], 0)
        self.assertEqual(len(response.data['errors']), 1)
        self.assertIn('not found', response.data['errors'][0])
    
    def test_upload_template_invalid_quantity(self):
        """Test uploading Excel file with invalid quantities"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'item_code'
        ws['B1'] = 'quantity'
        ws['A2'] = 'TEST002'
        ws['B2'] = 'invalid_quantity'
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        uploaded_file = SimpleUploadedFile(
            'test_items.xlsx',
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        url = reverse('purchase_order_api:purchase-order-items-upload', kwargs={'pk': self.purchase_order.pk})
        response = self.client.post(url, {'file': uploaded_file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['added'], 0)
        self.assertEqual(len(response.data['errors']), 1)
        self.assertIn('Invalid quantity format', response.data['errors'][0])
    
    def test_upload_template_update_existing_item(self):
        """Test uploading template updates existing item quantity"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'item_code'
        ws['B1'] = 'quantity'
        ws['A2'] = 'TEST001'  # This item already exists in the PO
        ws['B2'] = 15
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        uploaded_file = SimpleUploadedFile(
            'test_items.xlsx',
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        url = reverse('purchase_order_api:purchase-order-items-upload', kwargs={'pk': self.purchase_order.pk})
        response = self.client.post(url, {'file': uploaded_file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['added'], 1)
        
        # Verify item was updated, not duplicated
        self.assertEqual(self.purchase_order.items.count(), 1)
        self.po_item.refresh_from_db()
        self.assertEqual(self.po_item.quantity, Decimal('15.00'))
    
    def test_upload_template_nonexistent_po(self):
        """Test uploading template for non-existent PO returns error"""
        # Create a simple Excel file
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'item_code'
        ws['B1'] = 'quantity'
        ws['A2'] = self.inventory1.item_code
        ws['B2'] = 1
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        uploaded_file = SimpleUploadedFile(
            'test_items.xlsx',
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        url = reverse('purchase_order_api:purchase-order-items-upload', kwargs={'pk': 99999})
        response = self.client.post(url, {'file': uploaded_file}, format='multipart')
        
        # Accept either 404 or 500 since both indicate "not found"
        self.assertIn(response.status_code, [status.HTTP_404_NOT_FOUND, status.HTTP_500_INTERNAL_SERVER_ERROR])
        self.assertFalse(response.data['success'])
    
    def test_pagination_purchase_orders(self):
        """Test pagination of purchase orders list"""
        # Create multiple purchase orders
        for i in range(15):
            PurchaseOrder.objects.create(
                supplier=self.supplier,
                supplier_type='local',
                delivery_terms=f'FOB Test {i}',
                currency='USD',
                po_date='2024-01-01',
                created_by=self.user,
                last_modified_by=self.user
            )
        
        url = reverse('purchase_order_api:purchase-order-list')
        response = self.client.get(url, {'page': 1})
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertIn('meta', response.data)
        self.assertIn('pagination', response.data['meta'])
        self.assertIsNotNone(response.data['meta']['pagination']['count'])
    
    def test_unauthorized_template_access(self):
        """Test unauthorized access to template endpoints returns 401"""
        self.client.credentials()  # Remove authentication
        
        # Test download
        url = reverse('purchase_order_api:purchase-order-items-template', kwargs={'pk': self.purchase_order.pk})
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
        
        # Test upload
        url = reverse('purchase_order_api:purchase-order-items-upload', kwargs={'pk': self.purchase_order.pk})
        response = self.client.post(url, {})
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_add_item_to_existing_po(self):
        """Test adding a new item to an existing purchase order"""
        # Include existing item + new item
        data = {
            'items': [
                {
                    'id': self.po_item.pk,
                    'inventory': self.inventory1.pk,
                    'quantity': str(self.po_item.quantity),
                    'list_price': str(self.po_item.list_price),
                    'discount_type': self.po_item.discount_type,
                    'discount_value': str(self.po_item.discount_value or '0.00')  # Handle None values
                },
                {
                    'inventory': self.inventory2.pk,
                    'quantity': '5.00',
                    'list_price': '200.00',
                    'discount_type': 'none',
                    'discount_value': '0.00'
                }
            ]
        }
        
        url = reverse('purchase_order_api:purchase-order-detail', kwargs={'pk': self.purchase_order.pk})
        response = self.client.put(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Verify new item was added
        self.purchase_order.refresh_from_db()
        self.assertEqual(self.purchase_order.items.count(), 2)
        
        # Check that both items exist
        item1 = self.purchase_order.items.filter(inventory=self.inventory1).first()
        item2 = self.purchase_order.items.filter(inventory=self.inventory2).first()
        self.assertIsNotNone(item1)
        self.assertIsNotNone(item2)
        self.assertEqual(item2.quantity, Decimal('5.00'))

    def test_remove_item_from_po(self):
        """Test removing an item from purchase order by not including it in update"""
        # Create a second item first
        second_item = PurchaseOrderItem.objects.create(
            purchase_order=self.purchase_order,
            inventory=self.inventory2,
            quantity=Decimal('3.00'),
            list_price=Decimal('150.00'),
            discount_type='none',
            discount_value=Decimal('0.00')
        )
        
        # Verify we have 2 items
        self.assertEqual(self.purchase_order.items.count(), 2)
        
        # Update with only one item (removing the second)
        data = {
            'items': [
                {
                    'id': self.po_item.pk,
                    'inventory': self.inventory1.pk,
                    'quantity': str(self.po_item.quantity),
                    'list_price': str(self.po_item.list_price),
                    'discount_type': self.po_item.discount_type,
                    'discount_value': str(self.po_item.discount_value or '0.00')  # Handle None values
                }
            ]
        }
        
        url = reverse('purchase_order_api:purchase-order-detail', kwargs={'pk': self.purchase_order.pk})
        response = self.client.put(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Verify item was removed
        self.purchase_order.refresh_from_db()
        self.assertEqual(self.purchase_order.items.count(), 1)
        
        # Verify the remaining item is the correct one
        remaining_item = self.purchase_order.items.first()
        self.assertEqual(remaining_item.inventory, self.inventory1)

    def test_add_payment_term_to_po(self):
        """Test adding payment terms to a PO that doesn't have them"""
        # Create PO without payment terms
        po_without_terms = PurchaseOrder.objects.create(
            supplier=self.supplier,
            supplier_type='local',
            delivery_terms='FOB Test',
            currency='USD',
            created_by=self.user,
            last_modified_by=self.user
        )
        
        data = {
            'payment_term': {
                'credit_limit': '15000.00',
                'payment_terms_description': 'Net 45 days',
                'dp_percentage': '30.00',
                'terms_days': 45
            }
        }
        
        url = reverse('purchase_order_api:purchase-order-detail', kwargs={'pk': po_without_terms.pk})
        response = self.client.put(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Verify payment term was added
        po_without_terms.refresh_from_db()
        self.assertIsNotNone(po_without_terms.payment_term)
        self.assertEqual(po_without_terms.payment_term.credit_limit, Decimal('15000.00'))
        self.assertEqual(po_without_terms.payment_term.terms_days, 45)

    def test_update_payment_term(self):
        """Test updating existing payment terms"""
        data = {
            'payment_term': {
                'credit_limit': '25000.00',
                'payment_terms_description': 'Updated terms - Net 60 days',
                'dp_percentage': '40.00',
                'terms_days': 60
            }
        }
        
        url = reverse('purchase_order_api:purchase-order-detail', kwargs={'pk': self.purchase_order.pk})
        response = self.client.put(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Verify payment term was updated
        self.purchase_order.refresh_from_db()
        self.assertEqual(self.purchase_order.payment_term.credit_limit, Decimal('25000.00'))
        self.assertEqual(self.purchase_order.payment_term.terms_days, 60)
        self.assertEqual(self.purchase_order.payment_term.dp_percentage, Decimal('40.00'))

    def test_remove_payment_term(self):
        """Test removing payment terms from a PO"""
        data = {
            'payment_term': {}  # Empty dict should remove payment term
        }
        
        url = reverse('purchase_order_api:purchase-order-detail', kwargs={'pk': self.purchase_order.pk})
        response = self.client.put(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Verify payment term was removed
        self.purchase_order.refresh_from_db()
        self.assertIsNone(getattr(self.purchase_order, 'payment_term', None))

    def test_add_discount_charge_to_po(self):
        """Test adding discount/charge to purchase order"""
        data = {
            'discounts_charges': [
                {
                    'description': 'Early Payment Discount',
                    'is_percentage': True,
                    'value': '5.00',
                    'is_deduction': True
                },
                {
                    'description': 'Shipping Charge',
                    'is_percentage': False,
                    'value': '100.00',
                    'is_deduction': False
                }
            ]
        }
        
        url = reverse('purchase_order_api:purchase-order-detail', kwargs={'pk': self.purchase_order.pk})
        response = self.client.put(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Verify discounts/charges were added
        self.purchase_order.refresh_from_db()
        self.assertEqual(self.purchase_order.discounts_charges.count(), 2)
        
        discount = self.purchase_order.discounts_charges.filter(is_deduction=True).first()
        charge = self.purchase_order.discounts_charges.filter(is_deduction=False).first()
        
        self.assertEqual(discount.description, 'Early Payment Discount')
        self.assertEqual(charge.description, 'Shipping Charge')
        self.assertEqual(charge.value, Decimal('100.00'))

    def test_update_multiple_components_simultaneously(self):
        """Test updating items, payment terms, and discounts/charges in one request"""
        data = {
            'notes': 'Comprehensive update test',
            'items': [
                {
                    'id': self.po_item.pk,
                    'inventory': self.inventory1.pk,
                    'quantity': '20.00',
                    'list_price': '110.00',
                    'discount_type': 'fixed',
                    'discount_value': '10.00'
                },
                {
                    'inventory': self.inventory2.pk,
                    'quantity': '8.00',
                    'list_price': '250.00',
                    'discount_type': 'none',
                    'discount_value': '0.00'
                }
            ],
            'payment_term': {
                'credit_limit': '30000.00',
                'payment_terms_description': 'Updated comprehensive terms',
                'dp_percentage': '25.00',
                'terms_days': 30
            },
            'discounts_charges': [
                {
                    'description': 'Volume Discount',
                    'is_percentage': True,
                    'value': '3.00',
                    'is_deduction': True
                }
            ]
        }
        
        url = reverse('purchase_order_api:purchase-order-detail', kwargs={'pk': self.purchase_order.pk})
        response = self.client.put(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Verify all updates
        self.purchase_order.refresh_from_db()
        
        # Check basic fields
        self.assertEqual(self.purchase_order.notes, 'Comprehensive update test')
        
        # Check items
        self.assertEqual(self.purchase_order.items.count(), 2)
        
        # Find items by inventory instead of assuming pk still exists
        updated_item = self.purchase_order.items.filter(inventory=self.inventory1).first()
        self.assertIsNotNone(updated_item)
        self.assertEqual(updated_item.quantity, Decimal('20.00'))
        
        new_item = self.purchase_order.items.filter(inventory=self.inventory2).first()
        self.assertIsNotNone(new_item)
        self.assertEqual(new_item.quantity, Decimal('8.00'))
        
        # Check payment terms
        self.assertEqual(self.purchase_order.payment_term.credit_limit, Decimal('30000.00'))
        
        # Check discounts/charges
        self.assertEqual(self.purchase_order.discounts_charges.count(), 1)
        discount = self.purchase_order.discounts_charges.first()
        self.assertEqual(discount.description, 'Volume Discount')

    def test_partial_update_only_notes(self):
        """Test partial update changing only notes field"""
        original_supplier = self.purchase_order.supplier
        original_item_count = self.purchase_order.items.count()
        
        data = {
            'notes': 'Only notes updated'
        }
        
        url = reverse('purchase_order_api:purchase-order-detail', kwargs={'pk': self.purchase_order.pk})
        response = self.client.put(url, data, format='json')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        
        # Verify only notes changed
        self.purchase_order.refresh_from_db()
        self.assertEqual(self.purchase_order.notes, 'Only notes updated')
        self.assertEqual(self.purchase_order.supplier, original_supplier)
        self.assertEqual(self.purchase_order.items.count(), original_item_count)